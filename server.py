"""
13F Tracker - Local Server
==========================
Run:  python server.py
Open: http://localhost:8000

Fetches live 13F data from SEC EDGAR on each startup (cached for 24h).
No API keys, no accounts, no deployment needed.

If 13f.db is missing but 13f.sql.zip (or 13f.sql.gz) exists, the database
is automatically rebuilt from the compressed dump on first launch (~30-60s
one-time step).
"""

import http.server
import json
import os
import time
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET
import re
import threading
import gzip
import io
from pathlib import Path
from datetime import date, datetime, timedelta
import shutil


import math
import copy
import tempfile

# Excel export (openpyxl) — optional; xlsx download disabled if missing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# yfinance for price performance since filing date
try:
    import yfinance as yf
    _YF_AVAILABLE = True
except ImportError:
    _YF_AVAILABLE = False
    print("  [PERF] yfinance not installed — run: pip install yfinance")
    print("         Price performance will be unavailable until installed.")

# ─────────────────────────────────────────────
#  INVESTOR REGISTRY
#  To add a new investor, just append an entry.
#  CIK: find at https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company=NAME&type=13F
# ─────────────────────────────────────────────
INVESTORS = [
    {
        'id': 'buffett',
        'name': 'Warren Buffett',
        'firm': 'Berkshire Hathaway Inc',
        'strategy': 'Value / Long-term',
        'color': '#1549a8',
        'cik': '0001067983',
        'category': 'Investor / Company',
        'secName': 'BERKSHIRE HATHAWAY INC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'burry',
        'name': 'Michael Burry',
        'firm': 'Scion Asset Management LLC',
        'strategy': 'Value / Contrarian',
        'color': '#4c1d95',
        'cik': '0001649339',
        'category': 'Investor / Hedge Fund',
        'secName': 'SCION ASSET MANAGEMENT, LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'cathie_wood',
        'name': 'Cathie Wood',
        'firm': 'ARK Investment Management LLC',
        'strategy': 'Innovation / Growth',
        'color': '#ef4444',
        'cik': '0001697748',
        'category': 'Investor / ETF Adviser',
        'secName': 'ARK INVESTMENT MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'dalio',
        'name': 'Ray Dalio',
        'firm': 'Bridgewater Associates, LP',
        'strategy': 'Macro / Diversified',
        'color': '#0f766e',
        'cik': '0001350694',
        'category': 'Investor / Hedge Fund',
        'secName': 'BRIDGEWATER ASSOCIATES, LP',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'ackman',
        'name': 'Bill Ackman',
        'firm': 'Pershing Square Capital Management, L.P.',
        'strategy': 'Activist / Concentrated',
        'color': '#6d28d9',
        'cik': '0001336528',
        'category': 'Investor / Hedge Fund',
        'secName': 'PERSHING SQUARE CAPITAL MANAGEMENT, L.P.',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'tepper',
        'name': 'David Tepper',
        'firm': 'Appaloosa Management L.P.',
        'strategy': 'Macro / Opportunistic',
        'color': '#be185d',
        'cik': '0001656456',
        'category': 'Investor / Hedge Fund',
        'secName': 'APPALOOSA LP',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'druckenmiller',
        'name': 'Stanley Druckenmiller',
        'firm': 'Duquesne Family Office LLC',
        'strategy': 'Macro / Growth',
        'color': '#b45309',
        'cik': '0001536411',
        'category': 'Investor / Family Office',
        'secName': 'DUQUESNE FAMILY OFFICE LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'griffin',
        'name': 'Ken Griffin',
        'firm': 'Citadel Advisors LLC',
        'strategy': 'Multi-strategy',
        'color': '#0ea5e9',
        'cik': '0001423053',
        'category': 'Investor / Hedge Fund',
        'secName': 'CITADEL ADVISORS LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'cohen',
        'name': 'Steven Cohen',
        'firm': 'Point72 Asset Management, L.P.',
        'strategy': 'Long/short equity',
        'color': '#2563eb',
        'cik': '0001603466',
        'category': 'Investor / Hedge Fund',
        'secName': 'POINT72 ASSET MANAGEMENT, L.P.',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'klarman',
        'name': 'Seth Klarman',
        'firm': 'Baupost Group LLC/MA',
        'strategy': 'Value / Deep Discount',
        'color': '#9f3a00',
        'cik': '0001061768',
        'category': 'Investor / Hedge Fund',
        'secName': 'BAUPOST GROUP LLC/MA',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'einhorn',
        'name': 'David Einhorn',
        'firm': 'Greenlight Capital, Inc.',
        'strategy': 'Long/Short Value',
        'color': '#065f46',
        'cik': '0001079114',
        'category': 'Investor / Hedge Fund',
        'secName': 'GREENLIGHT CAPITAL INC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'marks',
        'name': 'Howard Marks',
        'firm': 'Oaktree Capital Management LP',
        'strategy': 'Credit / Value',
        'color': '#78350f',
        'cik': '0000949509',
        'category': 'Investor / Asset Manager',
        'secName': 'OAKTREE CAPITAL MANAGEMENT LP',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'dorsey',
        'name': 'Pat Dorsey',
        'firm': 'Dorsey Asset Management, LLC',
        'strategy': 'Quality / Moat',
        'color': '#0891b2',
        'cik': '0001671657',
        'category': 'Investor / Asset Manager',
        'secName': 'DORSEY ASSET MANAGEMENT, LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'greenblatt',
        'name': 'Joel Greenblatt',
        'firm': 'Gotham Asset Management, LLC',
        'strategy': 'Value / Quant',
        'color': '#7c3aed',
        'cik': '0001510387',
        'category': 'Investor / Asset Manager',
        'secName': 'GOTHAM ASSET MANAGEMENT, LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'hohn',
        'name': 'Chris Hohn',
        'firm': 'TCI Fund Management Ltd',
        'strategy': 'Concentrated / Activist',
        'color': '#dc2626',
        'cik': '0001647251',
        'category': 'Investor / Hedge Fund',
        'secName': 'TCI FUND MANAGEMENT LTD',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'halvorsen',
        'name': 'Andreas Halvorsen',
        'firm': 'Viking Global Investors LP',
        'strategy': 'Long/short equity',
        'color': '#16a34a',
        'cik': '0001103804',
        'category': 'Investor / Hedge Fund',
        'secName': 'VIKING GLOBAL INVESTORS LP',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'loeb',
        'name': 'Daniel Loeb',
        'firm': 'Third Point LLC',
        'strategy': 'Activist / Event-driven',
        'color': '#ea580c',
        'cik': '0001040273',
        'category': 'Investor / Hedge Fund',
        'secName': 'THIRD POINT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'gates',
        'name': 'Bill Gates',
        'firm': 'Gates Foundation Trust',
        'strategy': 'Foundation / Long-term',
        'color': '#15803d',
        'cik': '0001166559',
        'category': 'Foundation / Family',
        'secName': 'GATES FOUNDATION TRUST',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'akre',
        'name': 'Chuck Akre',
        'firm': 'Akre Capital Management LLC',
        'strategy': 'Quality compounders',
        'color': '#ca8a04',
        'cik': '0001112520',
        'category': 'Investor / Asset Manager',
        'secName': 'AKRE CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'gabelli',
        'name': 'Mario Gabelli',
        'firm': 'GAMCO Investors, Inc. et al',
        'strategy': 'Value / Private Market Value',
        'color': '#6366f1',
        'cik': '0000807249',
        'category': 'Investor / Asset Manager',
        'secName': 'GAMCO INVESTORS, INC. ET AL',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'watsa',
        'name': 'Prem Watsa',
        'firm': 'Fairfax Financial Holdings Ltd/Can',
        'strategy': 'Value / Insurance',
        'color': '#0d9488',
        'cik': '0000915191',
        'category': 'Investor / Company',
        'secName': 'FAIRFAX FINANCIAL HOLDINGS LTD/ CAN',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'icahn',
        'name': 'Carl Icahn',
        'firm': 'Icahn Carl C',
        'strategy': 'Activist',
        'color': '#991b1b',
        'cik': '0000921669',
        'category': 'Investor / Activist',
        'secName': 'ICAHN CARL C',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'soros',
        'name': 'George Soros',
        'firm': 'Soros Fund Management LLC',
        'strategy': 'Macro / Opportunistic',
        'color': '#1e40af',
        'cik': '0001029160',
        'category': 'Investor / Family Office',
        'secName': 'SOROS FUND MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'simons',
        'name': 'Jim Simons',
        'firm': 'Renaissance Technologies LLC',
        'strategy': 'Quantitative',
        'color': '#334155',
        'cik': '0001037389',
        'category': 'Investor / Hedge Fund',
        'secName': 'RENAISSANCE TECHNOLOGIES LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'tudor',
        'name': 'Paul Tudor Jones',
        'firm': 'Tudor Investment Corp et al',
        'strategy': 'Macro / Trading',
        'color': '#c2410c',
        'cik': '0000923093',
        'category': 'Investor / Hedge Fund',
        'secName': 'TUDOR INVESTMENT CORP ET AL',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'li_lu',
        'name': 'Li Lu',
        'firm': 'Himalaya Capital Management LLC',
        'strategy': 'Value / Concentrated',
        'color': '#047857',
        'cik': '0001709323',
        'category': 'Investor / Asset Manager',
        'secName': 'HIMALAYA CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'sosin',
        'name': 'Cliff Sosin',
        'firm': 'CAS Investment Partners, LLC',
        'strategy': 'Concentrated value',
        'color': '#7f1d1d',
        'cik': '0001697591',
        'category': 'Investor / Hedge Fund',
        'secName': 'CAS INVESTMENT PARTNERS, LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'tiger_global',
        'name': 'Tiger Global',
        'firm': 'Tiger Global Management LLC',
        'strategy': 'Growth / Tech',
        'color': '#f97316',
        'cik': '0001167483',
        'category': 'Hedge Fund',
        'secName': 'TIGER GLOBAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'millennium',
        'name': 'Millennium',
        'firm': 'Millennium Management LLC',
        'strategy': 'Multi-strategy',
        'color': '#0f172a',
        'cik': '0001273087',
        'category': 'Hedge Fund',
        'secName': 'MILLENNIUM MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'elliott',
        'name': 'Elliott Management',
        'firm': 'Elliott Investment Management L.P.',
        'strategy': 'Activist / Multi-strategy',
        'color': '#475569',
        'cik': '0001791786',
        'category': 'Hedge Fund',
        'secName': 'ELLIOTT INVESTMENT MANAGEMENT L.P.',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'lone_pine',
        'name': 'Lone Pine Capital',
        'firm': 'Lone Pine Capital LLC',
        'strategy': 'Growth / Long-short',
        'color': '#65a30d',
        'cik': '0001061165',
        'category': 'Hedge Fund',
        'secName': 'LONE PINE CAPITAL LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'deshaw',
        'name': 'D. E. Shaw',
        'firm': 'D. E. Shaw & Co., Inc.',
        'strategy': 'Quantitative / Multi-strategy',
        'color': '#0f766e',
        'cik': '0001009207',
        'category': 'Hedge Fund',
        'secName': 'D. E. SHAW & CO., INC.',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'coatue',
        'name': 'Coatue',
        'firm': 'Coatue Management LLC',
        'strategy': 'Technology / Growth',
        'color': '#2563eb',
        'cik': '0001135730',
        'category': 'Hedge Fund',
        'secName': 'COATUE MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'farallon',
        'name': 'Farallon Capital',
        'firm': 'Farallon Capital Management LLC',
        'strategy': 'Multi-strategy / Event-driven',
        'color': '#4d7c0f',
        'cik': '0000909661',
        'category': 'Hedge Fund',
        'secName': 'FARALLON CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'iconiq',
        'name': 'ICONIQ Capital',
        'firm': 'ICONIQ Capital, LLC',
        'strategy': 'Growth / Family Office',
        'color': '#0284c7',
        'cik': '0001585849',
        'category': 'Family Office / Adviser',
        'secName': 'ICONIQ CAPITAL, LLC',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'calpers',
        'name': 'CalPERS',
        'firm': 'California Public Employees Retirement System',
        'strategy': 'Public pension',
        'color': '#0369a1',
        'cik': '0000919079',
        'category': 'Public Pension',
        'secName': 'CALIFORNIA PUBLIC EMPLOYEES RETIREMENT SYSTEM',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
    {
        'id': 'rokos',
        'name': 'Chris Rokos',
        'firm': 'Rokos Capital Management LLP',
        'strategy': 'Macro / Global',
        'color': '#0369a1',
        'cik': '0001666335',
        'category': 'Hedge Fund',
        'secName': 'ROKOS CAPITAL MANAGEMENT LLP',
        'verified13F': True,
        'verificationNote': 'CIK 0001666335 confirmed on SEC EDGAR.',
    },
    {
        'id': 'd1_capital',
        'name': 'Dan Sundheim',
        'firm': 'D1 Capital Partners LP',
        'strategy': 'Long/Short Equity',
        'color': '#7c3aed',
        'cik': '0001747057',
        'category': 'Hedge Fund',
        'secName': 'D1 CAPITAL PARTNERS L.P.',
        'verified13F': True,
        'verificationNote': 'CIK 0001747057 confirmed on SEC EDGAR.',
    },
    {
        'id': 'caxton',
        'name': 'Andrew Law',
        'firm': 'Caxton Associates LP',
        'strategy': 'Macro / Global',
        'color': '#b45309',
        'cik': '0002051323',
        'category': 'Hedge Fund',
        'secName': 'CAXTON ASSOCIATES LLP',
        'verified13F': True,
        'verificationNote': 'CIK 0000764611 confirmed on SEC EDGAR.',
    },
    {
        'id': 'perceptive',
        'name': 'Joseph Edelman',
        'firm': 'Perceptive Advisors LLC',
        'strategy': 'Health Care / Life Sciences',
        'color': '#166534',
        'cik': '0001224962',
        'category': 'Hedge Fund',
        'secName': 'PERCEPTIVE ADVISORS LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001360803 confirmed on SEC EDGAR.',
    },
    {
        'id': 'discovery',
        'name': 'Robert Citrone',
        'firm': 'Discovery Capital Management LLC',
        'strategy': 'Macro / Global',
        'color': '#9a3412',
        'cik': '0001024732',
        'category': 'Hedge Fund',
        'secName': 'DISCOVERY CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001024732 confirmed on SEC EDGAR.',
    },
    {
        'id': 'jericho',
        'name': 'Josh Resnick',
        'firm': 'Jericho Capital Asset Management LP',
        'strategy': 'Technology / Growth',
        'color': '#0891b2',
        'cik': '0001525234',
        'category': 'Hedge Fund',
        'secName': 'JERICHO CAPITAL ASSET MANAGEMENT L.P.',
        'verified13F': True,
        'verificationNote': 'CIK 0001525234 confirmed on SEC EDGAR.',
    },
    {
        'id': 'balyasny',
        'name': 'Dmitry Balyasny',
        'firm': 'Balyasny Asset Management LLC',
        'strategy': 'Multi-Strategy',
        'color': '#4c1d95',
        'cik': '0001218710',
        'category': 'Hedge Fund',
        'secName': 'BALYASNY ASSET MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001218710 confirmed on SEC EDGAR.',
    },
    {
        'id': 'sachem_head',
        'name': 'Scott Ferguson',
        'firm': 'Sachem Head Capital Management LP',
        'strategy': 'Activist / Value',
        'color': '#be185d',
        'cik': '0001582090',
        'category': 'Hedge Fund',
        'secName': 'SACHEM HEAD CAPITAL MANAGEMENT LP',
        'verified13F': True,
        'verificationNote': 'CIK 0001582090 confirmed on SEC EDGAR.',
    },
    {
        'id': 'aqr',
        'name': 'Cliff Asness',
        'firm': 'AQR Capital Management LLC',
        'strategy': 'Quant / Multi-Strategy',
        'color': '#1d4ed8',
        'cik': '0001167557',
        'category': 'Hedge Fund',
        'secName': 'AQR CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001167557 confirmed on SEC EDGAR.',
    },
    {
        'id': 'schonfeld',
        'name': 'Steven Schonfeld',
        'firm': 'Schonfeld Strategic Advisors LLC',
        'strategy': 'Multi-Strategy / Quant',
        'color': '#065f46',
        'cik': '0001782983',
        'category': 'Hedge Fund',
        'secName': 'SCHONFELD STRATEGIC ADVISORS LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001782983 confirmed on SEC EDGAR.',
    },
    {
        'id': 'whale_rock',
        'name': 'Alex Sacerdote',
        'firm': 'Whale Rock Capital Management LLC',
        'strategy': 'Technology / Growth',
        'color': '#0d7d5f',
        'cik': '0001387322',
        'category': 'Hedge Fund',
        'secName': 'WHALE ROCK CAPITAL MANAGEMENT LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001387322 confirmed on SEC EDGAR.',
    },
    {
        'id': 'situational_awareness',
        'name': 'Leopold Aschenbrenner',
        'firm': 'Situational Awareness LP',
        'strategy': 'AI Infrastructure / Thematic',
        'color': '#7c3aed',
        'cik': '0002045724',
        'category': 'Hedge Fund',
        'secName': 'SITUATIONAL AWARENESS LP',
        'verified13F': True,
        'verificationNote': 'CIK 0002045724. Portfolio is heavily options-based — large put exposure against semiconductors reported as long positions in 13F.',
    },
    {
        'id': 'worldly_partners',
        'name': 'Arvind Navaratnam',
        'firm': 'Worldly Partners Management LLC',
        'strategy': 'Value / Concentrated',
        'color': '#0d7d5f',
        'cik': '0001904574',
        'category': 'Hedge Fund',
        'secName': 'WORLDLY PARTNERS MANAGEMENT, LLC',
        'verified13F': True,
        'verificationNote': 'CIK 0001904574 confirmed on SEC EDGAR.',
    },
    {
        'id': 'nys_common',
        'name': 'New York State Common Retirement Fund',
        'firm': 'New York State Common Retirement Fund',
        'strategy': 'Public pension',
        'color': '#1d4ed8',
        'cik': '0000810265',
        'category': 'Public Pension',
        'secName': 'NEW YORK STATE COMMON RETIREMENT FUND',
        'verified13F': True,
        'verificationNote': 'CIK selected for SEC 13F-HR institutional-manager filings; famous person names are mapped to their SEC reporting manager where applicable.',
    },
]

CACHE_DIR   = Path(__file__).parent / "cache"
CACHE_TTL   = 60 * 60 * 24   # 24 hours in seconds
SEC_HEADERS = {"User-Agent": "13FTracker Alejandro llamazaresalex148@gmail.com"}   # SEC requires a User-Agent

# ─────────────────────────────────────────────
#  SEC EDGAR HELPERS
# ─────────────────────────────────────────────
import time
LAST_SEC_CALL = 0
SEC_DELAY = .35

import math

def clean_json_value(obj):
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    if isinstance(obj, dict):
        return {k: clean_json_value(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [clean_json_value(v) for v in obj]

    return obj


def sec_get(url):
    """Fetch a URL from SEC EDGAR with throttling."""
    global LAST_SEC_CALL

    elapsed = time.time() - LAST_SEC_CALL
    if elapsed < SEC_DELAY:
        time.sleep(SEC_DELAY - elapsed)

    LAST_SEC_CALL = time.time()

    req = urllib.request.Request(url, headers=SEC_HEADERS)

    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return r.read().decode("utf-8", errors="replace")

    except Exception as e:
        print(f"  [SEC] fetch error {url}: {e}")
        return None


def get_13f_filings(cik):
    """Return 13F-HR filings for a CIK, sorted closest to today's date first.

    SEC's submissions endpoint is usually newest-first, but we do not rely on
    that. On every fresh fetch, this function checks today's date and chooses
    the most recent 13F filing date that is not in the future. This prevents
    an older/inactive cached-looking entry from driving the site headline.
    """
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    raw = sec_get(url)
    if not raw:
        return []

    data = json.loads(raw)
    filings = data.get("filings", {}).get("recent", {})
    forms   = filings.get("form", [])
    accnums = filings.get("accessionNumber", [])
    dates   = filings.get("filingDate", [])

    today = date.today()
    results = []

    for form, acc, filing_date in zip(forms, accnums, dates):
        # Accept every 13F *holdings* variant, including older electronic form
        # types (13F-E, plain 13F) used pre-2013. Skip 13F-NT notices.
        fu = (form or "").upper()
        if not fu.startswith("13F") or "NT" in fu:
            continue
        try:
            parsed_date = datetime.strptime(filing_date, "%Y-%m-%d").date()
        except Exception:
            parsed_date = date.min

        # Ignore impossible future filing dates if SEC data ever contains one.
        if parsed_date <= today:
            results.append({
                "accession": acc,
                "date": filing_date,
                "parsed_date": parsed_date.isoformat(),
            })

    results.sort(key=lambda f: f.get("parsed_date", "0000-00-00"), reverse=True)
    return results


def get_filing_index(cik, accession):
    """Fetch a filing's document index and return the best holdings-table URL.

    Post-2013 filings store the information table as XML; pre-2013 filings store
    it as a fixed-width text/SGML (.txt) document. We prefer XML when present and
    fall back to the legacy text document (and finally to the complete submission
    .txt) so that holdings can be read all the way back to the 1990s.
    """
    acc_clean = accession.replace("-", "")
    filing_base = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_clean}/"
    index_url = filing_base + "index.json"

    raw = sec_get(index_url)
    if not raw:
        # Even without an index we can still try the complete submission text.
        return f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}.txt"

    data = json.loads(raw)
    items = data.get("directory", {}).get("item", [])
    names = [item.get("name", "") for item in items]

    # ── 1. Prefer an XML information table (modern filings) ──────────────────
    xml_files = [n for n in names if n.lower().endswith(".xml")]
    preferred, fallback = [], []
    for name in xml_files:
        lower = name.lower()
        if "primary" in lower:   # primary_doc.xml is the cover, not the holdings
            continue
        if any(token in lower for token in ["infotable", "info_table", "form13f", "13f", "xslform13f"]):
            preferred.append(name)
        else:
            fallback.append(name)
    xml_candidates = preferred + fallback
    if xml_candidates:
        return filing_base + xml_candidates[0]

    # ── 2. Legacy text/SGML document (pre-2013 filings) ──────────────────────
    txt_files = [n for n in names if n.lower().endswith(".txt")]
    if txt_files:
        # Prefer a document that names itself as a 13F holdings report.
        scored = sorted(
            txt_files,
            key=lambda n: (
                "13f" in n.lower(),
                "hr" in n.lower(),
                not n.lower().endswith("-index.txt"),   # de-prioritise index files
            ),
            reverse=True,
        )
        return filing_base + scored[0]

    # ── 3. Last resort: the complete submission text file (always exists) ────
    #    e.g. .../edgar/data/1067983/0001193125-12-234582.txt
    return f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}.txt"


def parse_holdings_xml(xml_text):
    """Parse SEC 13F XML information table into list of holding dicts.

    This version is intentionally defensive because SEC 13F XML files are not
    perfectly consistent. Some filings contain namespace prefixes such as
    <ns1:infoTable> without declaring the prefix, which causes ElementTree to
    raise "unbound prefix". We normalize those prefixes before parsing.
    """

    def clean_sec_xml(text):
        text = text.strip()

        # Remove XML declaration.
        text = re.sub(r"<\?xml[^>]*\?>", "", text).strip()

        # Remove xmlns declarations with single or double quotes.
        text = re.sub(r"\s+xmlns(:[A-Za-z_][\w.-]*)?\s*=\s*\"[^\"]*\"", "", text)
        text = re.sub(r"\s+xmlns(:[A-Za-z_][\w.-]*)?\s*=\s*'[^']*'", "", text)

        # Remove namespace prefixes from tags:
        # <ns1:infoTable> -> <infoTable>
        # </ns1:infoTable> -> </infoTable>
        text = re.sub(r"<(/?)([A-Za-z_][\w.-]*):", r"<\1", text)

        # Remove namespace prefixes from attribute names too, just in case.
        text = re.sub(r"\s+([A-Za-z_][\w.-]*):([A-Za-z_][\w.-]*)=", r" \2=", text)

        return text

    def local_name(tag):
        # Handles both "{namespace}tag" and plain "tag".
        if "}" in tag:
            tag = tag.split("}", 1)[1]
        if ":" in tag:
            tag = tag.split(":", 1)[1]
        return tag.lower()

    def find_child_text(el, *names):
        wanted = {n.lower() for n in names}

        for child in el.iter():
            if child is el:
                continue

            child_name = local_name(child.tag)

            if child_name in wanted and child.text:
                return child.text.strip()

            # Handles nested structure:
            # <shrsOrPrnAmt><sshPrnamt>123</sshPrnamt><sshPrnamtType>SH</sshPrnamtType></shrsOrPrnAmt>
            if child_name == "shrsoforprnamt":
                for grandchild in child.iter():
                    grandchild_name = local_name(grandchild.tag)
                    if grandchild_name in wanted and grandchild.text:
                        return grandchild.text.strip()

        return ""

    holdings = []

    try:
        xml_clean = clean_sec_xml(xml_text)
        root = ET.fromstring(xml_clean)

        entries = [
            el for el in root.iter()
            if local_name(el.tag) == "infotable"
        ]

        if not entries:
            print("  [PARSE] No infoTable entries found")
            return []

        for entry in entries:
            name = find_child_text(entry, "nameOfIssuer")
            cusip = find_child_text(entry, "cusip")
            value = find_child_text(entry, "value")  # SEC reports this in $thousands
            shares = find_child_text(entry, "sshPrnamt")
            sh_type = find_child_text(entry, "sshPrnamtType")
            put_call = find_child_text(entry, "putCall")

            if not name or not value:
                continue

            # Include puts/calls in AUM but flag them so the UI can handle them
            is_option = bool(put_call)

            if not is_option and sh_type.upper() not in ("SH", "SHARES", ""):
                continue

            try:
                val_raw = int(float(value.replace(",", "")))
                shares_n = int(float(shares.replace(",", ""))) if shares else 0
            except ValueError:
                continue

            holdings.append({
                "name": name.title(),
                "cusip": cusip.upper() if cusip else "",
                "class": (find_child_text(entry, "titleOfClass") or "").strip(),
                "value": val_raw,
                "shares": shares_n,
                "putCall": put_call.upper() if put_call else None,
            })

        # ── Unit detection ───────────────────────────────────────────────────
        # SEC filings should use $thousands per spec, but most modern filings
        # use raw dollars. Detect which by checking the median position value:
        #
        #   median < 500,000  → filed in $thousands → divide by 1,000
        #   median >= 500,000 → filed in raw dollars → divide by 1,000,000
        #
        # A normal median position is $10M–$100M.
        #   In $thousands that's 10,000–100,000   (well below 500,000)
        #   In raw dollars that's 10,000,000+      (well above 500,000)
        if holdings:
            raw_values = sorted(h["value"] for h in holdings)
            median_raw = raw_values[len(raw_values) // 2]
            max_raw    = raw_values[-1]  # sorted list, last = maximum

            # Use the maximum value to detect units.
            # A $thousands filer's largest position (e.g. $5B fund top holding)
            # = $5B = raw 5,000,000. A raw-dollar filer's smallest top position
            # is easily $50M+ = raw 50,000,000. Threshold of 50M safely separates them.
            # Try $thousands first (SEC spec default), then raw dollars.
            # Pick whichever gives a plausible total AUM.
            raw_total = sum(h["value"] for h in holdings)
            aum_if_thousands  = raw_total / 1_000
            aum_if_rawdollars = raw_total / 1_000_000
            PLAUSIBLE_MIN = 0.1
            PLAUSIBLE_MAX = 50_000_000
            thou_ok = PLAUSIBLE_MIN <= aum_if_thousands  <= PLAUSIBLE_MAX
            raw_ok  = PLAUSIBLE_MIN <= aum_if_rawdollars <= PLAUSIBLE_MAX

            if thou_ok and not raw_ok:
                divisor = 1_000
            elif raw_ok and not thou_ok:
                divisor = 1_000_000
            elif thou_ok and raw_ok:
                median_raw = sorted(h["value"] for h in holdings)[len(holdings) // 2]
                divisor = 1_000 if median_raw < 500_000 else 1_000_000
            else:
                divisor = 1_000

            unit_label = "$thousands" if divisor == 1_000 else "raw dollars"
            print(f"  [PARSE] Detected unit: {unit_label} (median raw value: {median_raw:,})")

            for h in holdings:
                h["value"] = round(h["value"] / divisor, 2)

            total_m = sum(h["value"] for h in holdings)
            if total_m < 1:
                print(f"  [PARSE] WARNING: total after conversion is ${total_m:.2f}M — "
                      f"raw median {median_raw:,}. Manual check recommended.")
            elif total_m < 100:
                print(f"  [PARSE] NOTICE: total is ${total_m:.1f}M — "
                      f"small fund or possible unit mismatch. Raw median: {median_raw:,}.")

    except Exception as e:
        print(f"  [PARSE] XML parse error: {e}")

        # Helpful debug: print the start of the file so you can see whether SEC
        # returned HTML, an error page, or unexpected XML.
        preview = xml_text[:250].replace("\\n", " ")
        print(f"  [PARSE] File starts with: {preview}")

    return holdings


# ─────────────────────────────────────────────
#  LEGACY TEXT 13F PARSER (pre-2013 filings)
#  Before ~mid-2013 the SEC accepted 13F information tables as fixed-width
#  text/SGML documents (.txt) instead of XML. Those filings have no <infoTable>
#  elements, so parse_holdings_xml returns nothing for them. This parser reads
#  the fixed-width table directly so we can ingest holdings back to the 1990s.
# ─────────────────────────────────────────────

# A CUSIP is 9 chars (6 issuer + 2 issue + 1 numeric check digit). Newer text
# filings print it contiguously ("025816109"); older ones space the three
# segments ("025816 10 9"). Allow an optional single space between segments.
_TXT_CUSIP_CORE = r'[0-9A-Za-z]{6}\s?[0-9A-Za-z]{2}\s?[0-9]'

# A "full" data row introduces a security: name + title-of-class + CUSIP +
# Market Value + Shares/Principal. Everything left of the CUSIP is name+class.
_TXT_ROW_RE = re.compile(
    r'^(?P<pre>.*?)'
    r'\s+(?P<cusip>' + _TXT_CUSIP_CORE + r')'
    r'\s+(?P<value>[0-9][0-9,]*)'
    r'\s+(?P<shares>[0-9][0-9,]*)'
    r'(?P<rest>.*)$'
)

# A "lot" row is an *additional* position in the security named on a previous
# row: pre-2013 filings omit the name and CUSIP on repeats, leaving just the
# Value and Shares columns (indented, no issuer text in front).
_TXT_LOT_RE = re.compile(
    r'^(?P<lead>\s+)'
    r'(?P<value>[0-9][0-9,]*)'
    r'\s+(?P<shares>[0-9][0-9,]*)'
    r'(?P<rest>.*)$'
)

# Put / Call sometimes appears as its own column in these older filings.
_TXT_PUTCALL_RE = re.compile(r'\b(PUT|CALL)\b', re.IGNORECASE)

# Lines that must never be mistaken for a wrapped issuer-name fragment.
_TXT_NOISE_PREFIXES = (
    "TOTAL", "GRAND", "FORM 13F", "FORM 13-F", "NAME OF ISSUER", "TITLE",
    "CUSIP", "SHARES", "MARKET VALUE", "INVESTMENT", "VOTING", "NO.", "NONE",
    "CONFIDENTIAL", "REPORT", "PRINCIPAL", "OTHER MANAGERS", "PAGE", "COLUMN",
    "INFORMATION TABLE", "SUMMARY", "AMENDMENT", "<", "(THOUSANDS)",
)


def _txt_clean_name(raw):
    """Strip leader dots ('Co.............' -> 'Co') and collapse whitespace.
    Only runs of 2+ dots are removed, so abbreviations like 'U.S.' survive."""
    s = re.sub(r'\.{2,}', ' ', raw)
    s = re.sub(r'\s+', ' ', s).strip()
    return s.strip(' .,')


def _txt_is_rule_line(line):
    """A subtotal / separator rule made only of dashes, equals, dots, spaces."""
    s = line.strip()
    return bool(s) and bool(re.fullmatch(r'[-=_\s.]+', s))


def _txt_name_class_boundary(line):
    """If `line` is the dashes separator of the holdings table, return the
    column index where the Title-of-Class column begins (the start of the 2nd
    dash group). Otherwise return None.

    The holdings-table separator looks like:
        -------------------- ------------- --------- -------------- ...
    We require >=5 dash groups and a wide (>=10) first group so we don't latch
    onto the narrow separators used by the cover / summary tables.
    """
    if not line.strip() or not re.fullmatch(r'[-\s]+', line):
        return None
    groups = list(re.finditer(r'-+', line))
    if len(groups) < 5:
        return None
    if (groups[0].end() - groups[0].start()) < 10:
        return None
    return groups[1].start()


def _txt_looks_like_name_fragment(line):
    """True if a non-data line is plausibly part of a wrapped issuer name
    (e.g. 'American' / 'Express' sitting above 'Co... Com <cusip> ...')."""
    s = line.strip()
    if not s:
        return False
    if s.upper().startswith(_TXT_NOISE_PREFIXES):
        return False
    if not re.search(r'[A-Za-z]', s):
        return False
    if _txt_is_rule_line(s):
        return False
    if sum(c.isdigit() for c in s) > len(s) * 0.4:
        return False
    return len(s) <= 60


def parse_holdings_txt(text):
    """Parse a pre-2013 plain-text / SGML 13F information table into the same
    list-of-dicts shape that parse_holdings_xml returns.

    Column order in these filings is:
        Name of Issuer | Title of Class | CUSIP | Market Value (thousands)
        | Shares or Principal Amount | (discretion | managers | voting ...)

    Two layout quirks of these older filings are handled:
      * issuer names wrap over several lines, sometimes ending in leader dots;
      * repeated positions in one security omit the name and CUSIP, leaving a
        bare "value shares ..." row that inherits the security above it.
    """
    holdings = []
    pending_name = ""                       # wrapped-name fragments to prepend
    boundary = None                         # Name/Title split learned from rules
    cur_cusip = cur_name = cur_class = None # active security for bare lot rows

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n").replace("\t", "    ")
        stripped = line.strip()

        b = _txt_name_class_boundary(line)
        if b is not None:
            boundary = b
            pending_name = ""
            cur_cusip = cur_name = cur_class = None
            continue

        if not stripped or _txt_is_rule_line(line):
            pending_name = ""
            cur_cusip = cur_name = cur_class = None
            continue
        # Skip standalone SGML/markup tags: <PAGE>, <TABLE>, </TABLE>, <S>, <C> ...
        if stripped.startswith("<") and stripped.endswith(">"):
            continue

        # ── 1. A full row introduces a new security ──────────────────────────
        m = _TXT_ROW_RE.match(line)
        if m:
            cusip = re.sub(r'\s', '', m.group("cusip")).upper()
            if len(cusip) == 9:
                cusip_start = m.start("cusip")
                rest = m.group("rest") or ""

                # Separate issuer name from title-of-class. Prefer splitting on
                # runs of 2+ spaces (robust with leader-dot padding); only fall
                # back to the header column boundary when the split can't isolate
                # the class because the name fills the column with a single space
                # before it (e.g. "WELLS FARGO & CO NEW COM").
                chunks = [c for c in re.split(r'\s{2,}', m.group("pre").strip()) if c]
                if len(chunks) >= 2:
                    name_part, title_class = " ".join(chunks[:-1]), chunks[-1]
                elif boundary is not None and 0 < boundary < cusip_start:
                    name_part   = line[:boundary]
                    title_class = line[boundary:cusip_start].strip(" .")
                else:
                    name_part, title_class = (chunks[0] if chunks else ""), ""

                full_name = _txt_clean_name(
                    (pending_name + " " + name_part) if pending_name else name_part
                )
                pending_name = ""
                if not full_name:
                    continue

                try:
                    val_raw    = int(m.group("value").replace(",", ""))
                    shares_raw = int(m.group("shares").replace(",", ""))
                except ValueError:
                    continue

                pc = _TXT_PUTCALL_RE.search(title_class + " " + rest)
                cur_cusip, cur_name, cur_class = cusip, full_name, title_class
                holdings.append({
                    "name":    full_name.title(),
                    "cusip":   cusip,
                    "class":   title_class.strip(),
                    "value":   val_raw,
                    "shares":  shares_raw,
                    "putCall": pc.group(1).upper() if pc else None,
                })
                continue

        # ── 2. A bare lot row continues the current security ─────────────────
        if cur_cusip:
            lm = _TXT_LOT_RE.match(line)
            if lm and not re.search(r'[A-Za-z]', line[:lm.start("value")]):
                try:
                    val_raw    = int(lm.group("value").replace(",", ""))
                    shares_raw = int(lm.group("shares").replace(",", ""))
                except ValueError:
                    val_raw = None
                if val_raw is not None:
                    pc = _TXT_PUTCALL_RE.search(lm.group("rest") or "")
                    holdings.append({
                        "name":    cur_name.title(),
                        "cusip":   cur_cusip,
                        "class":   (cur_class or "").strip(),
                        "value":   val_raw,
                        "shares":  shares_raw,
                        "putCall": pc.group(1).upper() if pc else None,
                    })
                    continue

        # ── 3. Otherwise: a wrapped-name fragment, or noise ──────────────────
        if _txt_looks_like_name_fragment(line):
            pending_name = (pending_name + " " + stripped).strip() if pending_name else stripped
            cur_cusip = cur_name = cur_class = None   # a new issuer is starting
        else:
            pending_name = ""
            cur_cusip = cur_name = cur_class = None

    if not holdings:
        return []

    # ── Unit normalization ───────────────────────────────────────────────────
    # Legacy text 13F information tables ALWAYS report Market Value in thousands
    # (the column header literally reads "(In Thousands)"), so we convert a fixed
    # thousands→millions. We deliberately do NOT auto-detect thousands-vs-dollars
    # here the way the XML parser does: that heuristic relies on a median across
    # many positions and breaks on the tiny 1–3 holding confidential-treatment
    # amendments these filers submit — a single ~$1B position would otherwise be
    # misread as "raw dollars" and collapsed to ~$1M.
    for h in holdings:
        h["value"] = round(h["value"] / 1_000, 2)

    print(f"  [PARSE] legacy text 13F: {len(holdings)} holdings (values in $thousands)")
    return holdings


def parse_holdings(text):
    """Dispatch to the XML or the legacy-text parser based on file contents.

    Modern (post-2013) information tables contain <infoTable> elements; the old
    fixed-width text/SGML filings do not. Auto-detecting lets every caller stay
    format-agnostic.
    """
    if not text:
        return []
    if re.search(r'infotable', text, re.IGNORECASE):
        holdings = parse_holdings_xml(text)
        if holdings:
            return holdings
        # XML present but unparseable — fall through to the text parser, which
        # harmlessly returns [] if there's genuinely no fixed-width table.
    return parse_holdings_txt(text)


# ─────────────────────────────────────────────
#  DYNAMIC TICKER + SECTOR LOOKUP
#  Downloads SEC's master company file at startup.
#  Uses SIC codes for sector classification — no
#  hardcoded ticker lists needed.
# ─────────────────────────────────────────────

# SIC code ranges → sector name
# Source: SEC SIC manual (https://www.sec.gov/info/edgar/siccodes.htm)
SIC_SECTOR_MAP = [
    # Agriculture
    (100,   999,  "Agriculture"),
    # Mining
    (1000,  1499, "Energy"),
    # Oil & Gas
    (1300,  1399, "Energy"),
    # Construction
    (1500,  1799, "Industrials"),
    # Manufacturing — Food
    (2000,  2099, "Consumer Staples"),
    (2100,  2199, "Consumer Staples"),   # tobacco
    (2200,  2799, "Consumer Discr."),
    (2800,  2899, "Materials"),          # chemicals
    (2900,  2999, "Energy"),             # petroleum refining
    (3000,  3199, "Materials"),          # rubber, stone, glass
    (3200,  3299, "Materials"),
    (3300,  3399, "Materials"),          # metals
    (3400,  3499, "Industrials"),        # fabricated metal
    (3500,  3599, "Industrials"),        # industrial machinery
    (3600,  3699, "Technology"),         # electronic equipment
    (3670,  3679, "Technology"),         # semiconductors
    (3700,  3799, "Consumer Discr."),    # vehicles
    (3800,  3899, "Health Care"),        # instruments
    (3900,  3999, "Consumer Discr."),    # misc manufacturing
    # Transportation
    (4000,  4099, "Industrials"),
    (4100,  4299, "Industrials"),
    (4400,  4599, "Industrials"),
    (4600,  4699, "Energy"),             # pipelines
    (4700,  4799, "Industrials"),        # air transport
    (4800,  4899, "Communication"),      # telecom
    (4900,  4999, "Utilities"),
    # Wholesale
    (5000,  5199, "Industrials"),
    # Retail
    (5200,  5999, "Consumer Discr."),
    (5400,  5499, "Consumer Staples"),   # food stores
    (5900,  5999, "Consumer Discr."),
    # Finance & Insurance
    (6000,  6099, "Financials"),         # banks
    (6100,  6199, "Financials"),         # credit
    (6200,  6299, "Financials"),         # security brokers
    (6300,  6399, "Financials"),         # insurance
    (6400,  6411, "Financials"),
    (6500,  6599, "Real Estate"),
    (6700,  6799, "Financials"),         # holding companies
    # Services
    (7000,  7099, "Consumer Discr."),    # hotels
    (7200,  7299, "Consumer Discr."),    # laundry, beauty
    (7300,  7399, "Technology"),         # computer services
    (7370,  7379, "Technology"),         # computer programming
    (7380,  7389, "Industrials"),        # security services
    (7500,  7599, "Consumer Discr."),    # auto repair
    (7600,  7699, "Consumer Discr."),    # misc repair
    (7800,  7819, "Communication"),      # motion picture
    (7900,  7999, "Consumer Discr."),    # amusement
    (8000,  8099, "Health Care"),        # health services
    (8100,  8199, "Financials"),         # legal
    (8200,  8299, "Consumer Discr."),    # education
    (8300,  8399, "Consumer Discr."),    # social services
    (8700,  8799, "Technology"),         # engineering/management
    (8900,  8999, "Industrials"),
    # Public admin
    (9000,  9999, "Government"),
]

def sic_to_sector(sic):
    """Map a numeric SIC code to a sector string."""
    if not sic:
        return None
    try:
        s = int(sic)
    except (ValueError, TypeError):
        return None
    # Special cases first
    if 3674 <= s <= 3674: return "Technology"   # semiconductors
    if 7372 <= s <= 7372: return "Technology"   # prepackaged software
    if 6726 <= s <= 6726: return "Financials"   # investment offices (ETF-like)
    for lo, hi, sector in SIC_SECTOR_MAP:
        if lo <= s <= hi:
            return sector
    return None

# ── In-memory lookup tables (built at startup) ────────────────────────────
# cusip  → {"ticker": str, "name": str, "sic": int}
_CUSIP_MAP   = {}
# ticker → sector str
_TICKER_SECTOR = {}
# Whether the SEC master file has been loaded
_LOOKUP_LOADED = False
_LOOKUP_LOCK   = threading.Lock()


def load_sec_ticker_file():
    """
    Download SEC's company_tickers_exchange.json and build CUSIP→ticker map.
    Also download company_tickers.json (has CIK→ticker+SIC) for sector data.
    Both are cached for 7 days — they change slowly.
    """
    global _CUSIP_MAP, _TICKER_SECTOR, _LOOKUP_LOADED

    with _LOOKUP_LOCK:
        if _LOOKUP_LOADED:
            return

        ticker_cache = CACHE_DIR / "sec_tickers.json"
        sector_cache = CACHE_DIR / "sec_sectors.json"
        SEVEN_DAYS   = 7 * 24 * 3600

        # ── 1. CUSIP → ticker from company_tickers_exchange.json ──────────
        cusip_map = {}
        if ticker_cache.exists() and (time.time() - ticker_cache.stat().st_mtime) < SEVEN_DAYS:
            try:
                with open(ticker_cache) as f:
                    cusip_map = json.load(f)
                print(f"  [LOOKUP] Loaded {len(cusip_map):,} CUSIP→ticker entries from cache")
            except Exception:
                cusip_map = {}

        if not cusip_map:
            print("  [LOOKUP] Downloading SEC company tickers …")
            # Try multiple known-good URLs for the SEC ticker file
            raw = (sec_get("https://www.sec.gov/files/company_tickers_exchange.json") or
                   sec_get("https://www.sec.gov/files/company_tickers.json"))
            if raw:
                try:
                    data = json.loads(raw)
                    # Format: {"fields":["cik","name","ticker","exchange"], "data":[[...],...]}
                    fields = data.get("fields", [])
                    rows   = data.get("data", [])
                    # We need ticker. CUSIP is NOT in this file — we'll use CIK as bridge.
                    # Build CIK → ticker map here; CUSIP bridge done via submissions below.
                    cik_idx    = fields.index("cik")    if "cik"    in fields else None
                    ticker_idx = fields.index("ticker") if "ticker" in fields else None
                    name_idx   = fields.index("name")   if "name"   in fields else None
                    cik_to_ticker = {}
                    if cik_idx is not None and ticker_idx is not None:
                        for row in rows:
                            cik_to_ticker[str(row[cik_idx]).zfill(10)] = {
                                "ticker": row[ticker_idx],
                                "name":   row[name_idx] if name_idx is not None else "",
                            }
                    print(f"  [LOOKUP] Loaded {len(cik_to_ticker):,} CIK→ticker entries")
                    # Save CIK map temporarily — we'll build CUSIP map from CUSIP lookup below
                    cusip_map["_cik_to_ticker"] = cik_to_ticker
                    CACHE_DIR.mkdir(exist_ok=True)
                    with open(ticker_cache, "w") as f:
                        json.dump(cusip_map, f)
                except Exception as e:
                    print(f"  [LOOKUP] Parse error: {e}")

        # ── 2. Ticker → sector via SEC company_tickers.json (has SIC codes) ─
        ticker_sector = {}
        if sector_cache.exists() and (time.time() - sector_cache.stat().st_mtime) < SEVEN_DAYS:
            try:
                with open(sector_cache) as f:
                    ticker_sector = json.load(f)
                print(f"  [LOOKUP] Loaded {len(ticker_sector):,} ticker→sector entries from cache")
            except Exception:
                ticker_sector = {}

        if not ticker_sector:
            print("  [LOOKUP] Downloading SEC company names for sector keywords …")
            raw = (sec_get("https://www.sec.gov/files/company_tickers.json") or
                   sec_get("https://www.sec.gov/files/company_tickers_exchange.json"))
            if raw:
                try:
                    data = json.loads(raw)
                    # Format: {"0": {"cik_str":..,"ticker":..,"title":..}, ...}
                    # SIC is NOT here — we need submissions per company for SIC.
                    # But we can use name-based keyword matching as a fast path,
                    # and build a name→sector map that works for thousands of companies.
                    for entry in data.values():
                        ticker = entry.get("ticker", "")
                        name   = entry.get("title", "").lower()
                        if not ticker:
                            continue
                        sector = _sector_from_name(name)
                        if sector:
                            ticker_sector[ticker.upper()] = sector
                    print(f"  [LOOKUP] Built {len(ticker_sector):,} ticker→sector entries from names")
                    CACHE_DIR.mkdir(exist_ok=True)
                    with open(sector_cache, "w") as f:
                        json.dump(ticker_sector, f)
                except Exception as e:
                    print(f"  [LOOKUP] Sector parse error: {e}")

        # ── 3. Also fetch the CUSIP→ticker mapping file from SEC EDGAR ──────
        # SEC provides a CUSIP lookup via the full-text search index.
        # Better approach: use the company_tickers_exchange + manual overrides.
        cik_to_ticker = cusip_map.pop("_cik_to_ticker", {})

        _CUSIP_MAP.update(cusip_map)
        _TICKER_SECTOR.update(ticker_sector)
        # Store CIK→ticker for cross-reference
        _CUSIP_MAP["__cik_to_ticker__"] = cik_to_ticker
        _LOOKUP_LOADED = True
        print(f"  [LOOKUP] Ready: {len(_CUSIP_MAP):,} CUSIP entries, {len(_TICKER_SECTOR):,} sector entries")


def _sector_from_name(name_lower):
    """
    Fast keyword-based sector classification from company name.
    Used to build the initial ticker→sector cache.
    """
    n = name_lower
    if any(w in n for w in ["bancorp","bank","bancshares","financial","insurance","capital management",
                              "asset management","investment","savings","credit","mortgage","reit",
                              "trust co","holdings llc","fund"]):
        return "Financials"
    if any(w in n for w in ["pharma","biotech","therapeutics","biosciences","genomics",
                              "oncology","medical","health","hospital","clinic","surgical",
                              "diagnostic","drug","medicines"]):
        return "Health Care"
    if any(w in n for w in ["petroleum","oil","gas","energy","pipeline","drilling","refin",
                              "exploration","coal","mining","minerals","resources"]):
        return "Energy"
    if any(w in n for w in ["semiconductor","software","tech","systems","data","cloud",
                              "cyber","digital","computing","networks","internet","ai inc",
                              "artificial intelligence","semiconductor"]):
        return "Technology"
    if any(w in n for w in ["telecom","communications","wireless","cellular","broadband","fiber"]):
        return "Communication"
    if any(w in n for w in ["electric","utility","utilities","power","water","gas co"]):
        return "Utilities"
    if any(w in n for w in ["retail","consumer","foods","beverage","tobacco","household",
                              "personal care","grocery","supermarket","staples"]):
        return "Consumer Staples"
    if any(w in n for w in ["restaurant","hotel","resort","automotive","apparel","luxury",
                              "entertainment","media","leisure","gaming","cruise"]):
        return "Consumer Discr."
    if any(w in n for w in ["aerospace","defense","construction","engineering","transport",
                              "logistics","manufacturing","industrial","machinery","railroad",
                              "airline","shipping"]):
        return "Industrials"
    if any(w in n for w in ["chemical","materials","steel","aluminum","copper","gold",
                              "silver","mining","timber","packaging"]):
        return "Materials"
    if any(w in n for w in ["real estate","property","realty","reit","apartments","office park"]):
        return "Real Estate"
    if any(w in n for w in ["etf","index fund","ishares","vanguard","spdr","invesco"]):
        return "ETF"
    return None


# Hardcoded overrides for the most important tickers — these take priority
# over any dynamically derived value. Keep this SHORT — only major ones.
TICKER_OVERRIDES = {
    "AAPL":"Technology",   "MSFT":"Technology",   "NVDA":"Technology",
    "GOOGL":"Technology",  "GOOG":"Technology",   "META":"Technology",
    "AMZN":"Technology",   "TSLA":"Consumer Discr.", "ORCL":"Technology",
    "CRM":"Technology",    "INTC":"Technology",   "AMD":"Technology",
    "QCOM":"Technology",   "AVGO":"Technology",   "TSM":"Technology",
    "BIDU":"Technology",   "BABA":"Technology",   "JD":"Technology",
    "PDD":"Technology",    "NFLX":"Communication","CSCO":"Technology",
    "UBER":"Technology",   "BAC":"Financials",    "JPM":"Financials",
    "GS":"Financials",     "MS":"Financials",     "WFC":"Financials",
    "C":"Financials",      "BX":"Financials",     "V":"Financials",
    "MA":"Financials",     "AXP":"Financials",    "MCO":"Financials",
    "SPGI":"Financials",   "SCHW":"Financials",   "BN":"Financials",
    "KO":"Consumer Staples","PG":"Consumer Staples","PM":"Consumer Staples",
    "KHC":"Consumer Staples","KR":"Consumer Staples","WMT":"Consumer Staples",
    "COST":"Consumer Staples","MO":"Consumer Staples",
    "UNH":"Health Care",   "LLY":"Health Care",   "JNJ":"Health Care",
    "MRK":"Health Care",   "AMGN":"Health Care",  "CVS":"Health Care",
    "HCA":"Health Care",   "ELV":"Health Care",   "ABT":"Health Care",
    "CVX":"Energy",        "XOM":"Energy",        "OXY":"Energy",
    "COP":"Energy",        "SLB":"Energy",        "EOG":"Energy",
    "GE":"Industrials",    "DAL":"Industrials",   "HON":"Industrials",
    "CAT":"Industrials",   "RTX":"Industrials",   "UNP":"Industrials",
    "VZ":"Communication",  "T":"Communication",   "DIS":"Communication",
    "CMCSA":"Communication","CHTR":"Communication",
    "NEE":"Utilities",     "SO":"Utilities",      "DUK":"Utilities",
    "NRG":"Utilities",     "VST":"Utilities",
    "SPY":"ETF",           "QQQ":"ETF",           "IVV":"ETF",
    "IWM":"ETF",           "GLD":"ETF",           "GDX":"ETF",
    "EEM":"ETF",           "VWO":"ETF",
    "AMT":"Real Estate",   "PLD":"Real Estate",   "SPG":"Real Estate",
    "INVH":"Real Estate",  "GRBK":"Real Estate",
    "DOW":"Materials",     "LIN":"Materials",     "FCX":"Materials",
}


def _is_preferred_ticker(t):
    """True if a ticker symbol denotes a preferred / warrant / rights line
    rather than common stock — e.g. 'C-PR', 'ALL-PJ', 'MS-PQ', 'HPE-PC',
    'CELG-RI'. Keyed on a separator (-, ., ^, $, /) followed by P/W/R.
    Dual-class commons ('BRK.B', 'BRK-B', 'LGF-A') are NOT matched."""
    if not t:
        return False
    return bool(re.search(r'[-.\^$/](P|W|R)[A-Z0-9]*$', t.upper()))

# Title-of-class tokens that mean the line is NOT common stock.
_NONCOMMON_CLASS_RE = re.compile(
    r'\b(PFD|PREF|PREFERRED|DEPOSITARY|DEP\s+SHS|DEP|WARRANT|WT|WTS|RIGHT|RTS|'
    r'NOTE|NT|BOND|BD|DEB|CONVERTIBLE|CONV|UNIT|UNT)\b', re.I)

def _is_noncommon_class(title_class):
    """True if the 13F 'title of class' marks the holding as preferred,
    a warrant, a right, a note/bond, a unit, etc. (anything but common)."""
    if not title_class:
        return False
    return bool(_NONCOMMON_CLASS_RE.search(title_class))


def resolve_ticker(cusip, name, title_class=""):
    """
    Map a CUSIP to a ticker symbol.
    Strategy:
      0. If the title-of-class marks this as non-common (preferred / warrant /
         right / note / unit), flag it as excluded so it is never mispriced as
         the common stock.
      1. Static overrides (tiny, only for ambiguous CUSIPs)
      2. SEC company_tickers_exchange dynamic map (CIK bridge via name match),
         skipping preferred/warrant tickers so e.g. Citigroup common does not
         resolve to a 'C-PR' preferred symbol.
      3. Derive from company name as last resort
    """
    if not _LOOKUP_LOADED:
        load_sec_ticker_file()

    # 0. Genuine non-common securities: flag and exclude (kept out of pricing).
    #    The CUSIP keeps each distinct preferred/warrant line separate.
    if _is_noncommon_class(title_class):
        return "~PFD:" + (cusip[:6] if cusip else (name[:4].upper() if name else "???"))

    # Some CUSIPs are known to be problematic — direct map
    CUSIP_DIRECT = {
        # ── US Large Cap ──────────────────────────────────────────────────
        "037833100":"AAPL",  "023135106":"AMZN",  "02079K305":"GOOGL",
        "02079K107":"GOOG",  "594918104":"MSFT",  "67066G104":"NVDA",
        "30303M102":"META",  "88160R101":"TSLA",  "025816109":"AXP",
        "191216100":"KO",    "060505104":"BAC",   "166764100":"CVX",
        "50076Q106":"KHC",   "615369105":"MCO",   "254687106":"DVA",
        "499580100":"KR",    "92826C839":"V",     "57636Q104":"MA",
        "247361702":"DAL",   "742718109":"PG",    "718172109":"PM",
        "594901100":"MRK",   "126650100":"CVS",   "90353T100":"UBER",
        "097023105":"BX",    "98421M106":"BABA",  "72352L106":"PDD",
        "874039100":"TSM",   "595017104":"MU",    "539830103":"LLY",
        "65339F101":"NFLX",  "68389X105":"ORCL",  "458140100":"INTC",
        "808513105":"SCHW",  "031162100":"AMGN",  "46625H100":"JPM",
        "172967424":"C",     "91324P102":"UNH",   "26875P101":"ELV",
        "369604103":"GE",    "464287655":"IVV",   "91282C857":"QQQ",
        "78462F103":"SPY",   "464287804":"IWM",
        "76131D103":"QSR",   "43300A203":"HLT",
        "95040Q104":"WEN",   "741503207":"PNR",
        "45167R104":"INVH",  "G4124G107":"GRBK",  "12653C108":"CNX",
        "339750101":"FCNCA", "740260108":"HPQ",   "690879102":"OVV",
        "649464100":"NYCB",  "63938C108":"NRG",   "92840M102":"VST",
        "109696104":"BIDU",  "47215P106":"JD",
        "58155Q103":"MCK",   "740816101":"PNC",   "16119P108":"CHTR",
        "78463V107":"SPGI",  "585055106":"MET",   "921937106":"VZ",
        "00206R102":"T",     "29081T107":"EFX",   "316888203":"FMC",
        "500255104":"KMI",   "09062X103":"BIIB",
        "38141G104":"GOOGL", "049835107":"AVGO",
        "084670702":"BRK.B", "857477103":"STT",   "717081103":"PFE",
        "49460W201":"LPLA",  "40424L109":"GXO",
        "74965L101":"QRTEA", "423074103":"HPE",
        "78409V104":"CRM",   "53307C102":"LYFT",
        "70450Y103":"PYPL",  "14040H105":"CARR",  "92343V104":"VRT",
        "60855R100":"MOH",   "583191104":"MEDP",
        "531229854":"LULU",  "553530106":"MCK",   "009728106":"AIZ",
        "50540R409":"LHX",   "912093901":"GLBE",
        # ── CORRECTED: previously wrong CUSIPs ───────────────────────────
        "674599105":"OXY",   # Occidental Petroleum (correct CUSIP)
        "693718108":"OXY",   # Occidental Petroleum (alternate CUSIP)
        "20605P101":"CB",    # Chubb Corp (US entity)
        "H1467J104":"CB",    # Chubb Ltd (Swiss-domiciled, used in recent filings)
        "G16840T102":"CB",   # Chubb Ltd another variant
        # ── Foreign / Cayman CUSIPs (start with letters) ─────────────────
        "G16612109":"BN",    "G1890L107":"CP",    "G50871105":"JHG",
        "G21178105":"CRH",   "G0750C108":"AON",   "G1151C101":"BN",
        "G01460113":"IAC",   "G76225104":"RYAAY", "L02382106":"ACGL",
        "L5769L101":"MKL",   "53217V109":"LI",    "44107P104":"HPE",
        "20826N105":"COUP",  "11135F101":"BRKR",  "82811H100":"SHLD",
        "74975P104":"RCM",   "532457108":"ELI",
        # ── Additional large caps ─────────────────────────────────────────
        "023135106":"AMZN",  "46625H100":"JPM",   "172967424":"C",
        "742718109":"PG",    "079879103":"BBY",   "125523100":"CI",
        "22160K105":"COP",   "219350105":"COR",   "247361702":"DAL",
        "268648102":"EMR",   "292505104":"EQT",   "337932107":"FIS",
        "369550108":"GS",    "382550101":"GPC",   "413875105":"HAL",
        "441300100":"HON",   "44107P104":"HPE",   "458140100":"INTC",
        "46434V102":"ISRG",  "48666K109":"KKR",   "52465W106":"LIN",
        "571748102":"MAR",   "594918104":"MSFT",  "617478103":"MOS",
        "63888U108":"NKE",   "652482100":"NOC",   "670346105":"NUE",
        "693506107":"PLD",   "719413100":"PNC",   "723254106":"PXD",
        "731011105":"PKI",   "742718109":"PG",    "764038106":"RHI",
        "808513105":"SCHW",  "832696405":"SLB",   "842587107":"SPG",
        "872589107":"TGT",   "879585109":"TXN",   "88160R101":"TSLA",
        "900890107":"UAL",   "911312106":"UPS",   "913017109":"USB",
        "920148108":"VLO",   "924717100":"VZ",    "929042109":"WMT",
        "949746101":"WFC",   "98872P103":"ZM",    "025816109":"AXP",
    }

    if cusip in CUSIP_DIRECT:
        return CUSIP_DIRECT[cusip]

    # Try the dynamic CIK→ticker map via name matching
    cik_to_ticker = _CUSIP_MAP.get("__cik_to_ticker__", {})
    if cik_to_ticker and name:
        name_norm = name.strip().upper()
        for cik_padded, entry in cik_to_ticker.items():
            sec_name = entry.get("name", "").strip().upper()
            if sec_name and (sec_name == name_norm or name_norm.startswith(sec_name[:12]) or sec_name.startswith(name_norm[:12])):
                cand = entry["ticker"]
                # The CIK→ticker map is last-write-wins per company, so a
                # preferred/warrant listing can stand in for the whole issuer.
                # Skip those so common holdings fall through to the name table
                # (e.g. Citigroup -> "C" rather than "C-PR").
                if _is_preferred_ticker(cand):
                    continue
                return cand

    # Name → ticker lookup table (handles common cases where CUSIP map misses)
    # Keyed on the start of the company name as it appears in SEC filings (uppercase)
    NAME_TO_TICKER = {
        # Berkshire holdings
        "OCCIDENTAL":       "OXY",    "CHUBB":            "CB",
        "KRAFT HEINZ":      "KHC",    "MOODYS":           "MCO",
        "DAVITA":           "DVA",    "KROGER":           "KR",
        "CHEVRON":          "CVX",    "COCA-COLA":        "KO",
        "COCA COLA":        "KO",     "BANK OF AMERICA":  "BAC",
        "AMERICAN EXPRESS": "AXP",    "APPLE":            "AAPL",
        # Common holdings across funds
        "MICROSOFT":        "MSFT",   "ALPHABET":         "GOOGL",
        "AMAZON":           "AMZN",   "NVIDIA":           "NVDA",
        "META PLATFORMS":   "META",   "TESLA":            "TSLA",
        "BROADCOM":         "AVGO",   "UNITEDHEALTH":     "UNH",
        "JOHNSON":          "JNJ",    "JPMORGAN":         "JPM",
        "VISA":             "V",      "MASTERCARD":       "MA",
        "EXXON":            "XOM",    "WALMART":          "WMT",
        "PROCTER":          "PG",     "HOME DEPOT":       "HD",
        "ABBVIE":           "ABBV",   "MERCK":            "MRK",
        "COSTCO":           "COST",   "ABBOTT":           "ABT",
        "THERMO FISHER":    "TMO",    "PEPSICO":          "PEP",
        "SALESFORCE":       "CRM",    "ORACLE":           "ORCL",
        "ACCENTURE":        "ACN",    "CISCO":            "CSCO",
        "QUALCOMM":         "QCOM",   "APPLIED MATERIALS":"AMAT",
        "AMD":              "AMD",    "INTEL":            "INTC",
        "TEXAS INSTRUMENTS":"TXN",    "LINDE":            "LIN",
        "CATERPILLAR":      "CAT",    "HONEYWELL":        "HON",
        "UNION PACIFIC":    "UNP",    "RAYTHEON":         "RTX",
        "LOCKHEED":         "LMT",    "BOEING":           "BA",
        "DEERE":            "DE",     "GENERAL ELECTRIC": "GE",
        "3M":               "MMM",    "UNITED PARCEL":    "UPS",
        "FEDEX":            "FDX",    "NORFOLK":          "NSC",
        "DELTA AIR":        "DAL",    "SOUTHWEST":        "LUV",
        "AMERICAN AIRLINES":"AAL",    "UNITED AIRLINES":  "UAL",
        "MCDONALDS":        "MCD",    "STARBUCKS":        "SBUX",
        "CHIPOTLE":         "CMG",    "YUM":              "YUM",
        "NIKE":             "NKE",    "LULULEMON":        "LULU",
        "TJXCOS":           "TJX",    "ROSS STORES":      "ROST",
        "TARGET":           "TGT",    "DOLLAR GENERAL":   "DG",
        "DOLLAR TREE":      "DLTR",   "LOWES":            "LOW",
        "BEST BUY":         "BBY",    "AUTOZONE":         "AZO",
        "NETFLIX":          "NFLX",   "DISNEY":           "DIS",
        "COMCAST":          "CMCSA",  "CHARTER":          "CHTR",
        "VERIZON":          "VZ",     "ATT":              "T",
        "T-MOBILE":         "TMUS",   "WARNER BROS":      "WBD",
        "WARNER BROTHERS":  "WBD",    "FOX":              "FOX",
        "PARAMOUNT":        "PARA",   "SPOTIFY":          "SPOT",
        "PFIZER":           "PFE",    "ELI LILLY":        "LLY",
        "BRISTOL":          "BMY",    "AMGEN":            "AMGN",
        "GILEAD":           "GILD",   "REGENERON":        "REGN",
        "BIOGEN":           "BIIB",   "VERTEX":           "VRTX",
        "IQVIA":            "IQV",    "DANAHER":          "DHR",
        "MEDTRONIC":        "MDT",    "STRYKER":          "SYK",
        "EDWARDS":          "EW",     "INTUITIVE":        "ISRG",
        "UNITEDHEALTH":     "UNH",    "ELEVANCE":         "ELV",
        "CIGNA":            "CI",     "CENTENE":          "CNC",
        "HUMANA":           "HUM",    "CVS HEALTH":       "CVS",
        "MCKESSON":         "MCK",    "CARDINAL":         "CAH",
        "WELLSFARGO":       "WFC",    "WELLS FARGO":      "WFC",
        "CITIGROUP":        "C",      "GOLDMAN":          "GS",
        "MORGAN STANLEY":   "MS",     "BLACKSTONE":       "BX",
        "KKR":              "KKR",    "APOLLO":           "APO",
        "ARES":             "ARES",   "BERKSHIRE":        "BRK.B",
        "BLACKROCK":        "BLK",    "PIMCO":            "PIMCO",
        "SCHWAB":           "SCHW",   "FIDELITY":        "FNF",
        "INTERCONTINENTAL": "ICE",    "CME GROUP":        "CME",
        "MOODY":            "MCO",    "S&P GLOBAL":       "SPGI",
        "MSCI":             "MSCI",   "FACTSET":          "FDS",
        "AMERICAN TOWER":   "AMT",    "PROLOGIS":         "PLD",
        "CROWN CASTLE":     "CCI",    "SIMON PROPERTY":   "SPG",
        "EQUINIX":          "EQIX",   "DIGITAL REALTY":   "DLR",
        "CONOCOPHILLIPS":   "COP",    "PIONEER":          "PXD",
        "SCHLUMBERGER":     "SLB",    "HALLIBURTON":      "HAL",
        "BAKER HUGHES":     "BKR",    "WILLIAMS":         "WMB",
        "KINDER MORGAN":    "KMI",    "MARATHON":         "MPC",
        "VALERO":           "VLO",    "PHILLIPS 66":      "PSX",
        "NEXTERA":          "NEE",    "DUKE ENERGY":      "DUK",
        "SOUTHERN":         "SO",     "DOMINION":         "D",
        "AMERICAN ELECTRIC":"AEP",    "XCEL":             "XEL",
        "NRG ENERGY":       "NRG",    "VISTRA":           "VST",
        "COREWEAVE":        "CRWV",   "PPL":              "PPL",
        "LOCKHEED MARTIN":  "LMT",    "VENTURE GLOBAL":   "VG",
        "TAIWAN SEMI":      "TSM",    "SAMSUNG":          "SSNLF",
        "ASML":             "ASML",   "LUMENTUM":         "LITE",
        "SEA LIMITED":      "SE",     "SEA LTD":          "SE",
        "NATERA":           "NTRA",   "INSMED":           "INSM",
        "NEWAMSTERDAM":     "NAMS",   "STMICRO":          "STM",
        "WOODWARD":         "WWD",    "TEVA":             "TEVA",
        "ROKU":             "ROKU",   "COUPANG":          "CPNG",
        "CRH":              "CRH",    "FIGURE TECH":      "FGTH",
        "GLOBAL X":         "GCOW",   "YPF":              "YPF",
        "BBB FOODS":        "TBBB",   "ALCOA":            "AA",
        "OPTION CARE":      "OPCH",   "HDFC":             "HDB",
        # ── Historical / acquired companies (consensus top-15) ───────────
        # These traded under now-defunct tickers. yfinance still has their
        # historical price data, which is what the backtest needs.
        "GILLETTE":         "G",      "GILLETE":          "G",       # → P&G 2005
        "XYLAN":            "XYLN",                                  # → Alcatel 1999
        "CENDANT":          "CD",                                    # split 2006
        "OMEGA WORLDWIDE":  "OWW",                                   # healthcare REIT
        "LUCENT":           "LU",                                    # → Alcatel-Lucent 2006
        "MCI WORLDCOM":     "WCOM",   "WORLDCOM":         "WCOM",   # bankrupt 2002
        "JDS UNIPHASE":     "JDSU",                                  # now Viavi (VIAV)
        "AGRIBRANDS":       "AGX",                                   # → Cargill 2001
        "AMFM":             "AFM",                                   # → Clear Channel 2000
        "WARNER LAMBERT":   "WLA",                                   # → Pfizer 2000
        "ATLANTIC RICHFIELD":"ARC",                                  # → BP 2000
        "U S WEST":         "USW",    "US WEST":          "USW",    # → Qwest 2000
        "SBC COMMUN":       "SBC",                                   # → AT&T 2005
        "BESTFOOD":         "BFO",                                   # → Unilever 2000
        "NORTEL":           "NT",                                    # bankrupt 2009
        "PHARMACIA":        "PHA",                                   # → Pfizer 2003
        "XO COMMUN":        "XOCM",                                 # → Verizon
        "VOICESTREAM":      "VSTR",                                  # → T-Mobile 2001
        "AOL TIME WARNER":  "AOL",    "AOL":              "AOL",    # renamed Time Warner
        "ECHOSTAR":         "DISH",                                  # became DISH Network
        "DELL COMPUTER":    "DELL",   "DELL":             "DELL",
        "TYCO INTL":        "TYC",    "TYCO INT":         "TYC",
        "CYPRESS SEMI":     "CY",     "CYPRESS SEMICONDUCTOR": "CY",
        "FIRST DATA":       "FDC",                                   # → Fiserv 2019
        "KMART":            "KMRT",                                  # → Sears (SHLD)
        "GUIDANT":          "GDT",                                   # → Boston Scientific 2006
        "MONSANTO":         "MON",                                   # → Bayer 2018
        "WELLPOINT":        "WLP",                                   # now Elevance (ELV)
        "PRICELINE":        "PCLN",                                  # now Booking (BKNG)
        "WYETH":            "WYE",                                   # → Pfizer 2009
        "KRAFT FOOD":       "KFT",                                   # pre-split Kraft Foods
        "DIRECTV":          "DTV",    "DIRECT TV":        "DTV",    # → AT&T 2015
        "GENZYME":          "GENZ",                                  # → Sanofi 2011
        "EXPRESS SCRIPT":   "ESRX",                                  # → Cigna 2018
        "VIRGIN MEDIA":     "VMED",                                  # → Liberty Global
        "CBS CORP":         "CBS",    "CBS":              "CBS",
        "ALLERGAN":         "AGN",                                   # → AbbVie 2020
        "AVAGO":            "AVGO",                                  # merged into Broadcom
        "TIME WARNER":      "TWX",                                   # → AT&T 2018
        "ACTIVISION":       "ATVI",                                  # → Microsoft 2023
        "ALTABA":           "AABA",                                  # liquidated 2019
        "MOTOROLA":         "MOT",                                   # split MSI + MMI
        "VIACOM":           "VIA",                                   # merged w/ CBS
        "YAHOO":            "YHOO",                                  # → Altaba / Verizon
        "TECH DATA":        "TECD",                                  # → TD SYNNEX
    }

    # Try name lookup — check if company name starts with any known key
    name_upper = name.strip().upper()
    # Remove common suffixes for matching
    for suffix in [" INC", " CORP", " CO", " LTD", " LLC", " PLC", " NV", " SA",
                   " AG", " SE", " GROUP", " HOLDINGS", " COMPANY", " LIMITED"]:
        name_upper = name_upper.replace(suffix, "")
    name_upper = name_upper.strip()

    for key, ticker in NAME_TO_TICKER.items():
        if name_upper.startswith(key) or key in name_upper:
            return ticker

    # Absolute last resort: use CUSIP as the ticker key to prevent collisions.
    # We prefix with "~" so these never match real tickers in any map.
    # This means unknown holdings won't pollute well-known tickers like CB, AXP, MSFT
    # just because their company name starts with the same letters.
    if cusip and len(cusip) >= 6:
        return "~" + cusip[:6]
    words = re.sub(r"[^a-zA-Z0-9\s]", "", name).split()
    if not words:
        return "~???"
    first = words[0].upper()
    for suffix in ["INC", "CORP", "CO", "LTD", "LLC", "PLC", "THE", "NEW"]:
        if first == suffix and len(words) > 1:
            first = words[1].upper()
            break
    # Append first 4 chars of CUSIP to make it unique and non-colliding
    cusip_suffix = cusip[:4] if cusip else "????"
    return "~" + first[:4] + cusip_suffix[:2]


# Read-time correction for DBs ingested before the resolver fix.
# Existing rows may carry a contaminated preferred ticker (e.g. "C-PR" for a
# Citigroup COMMON holding). We re-resolve those from the stored CUSIP+name so
# the common ticker is recovered ("C") and priced split-adjusted by yfinance.
# Rows that re-resolve to another preferred/placeholder are left flagged.
# NOTE: the DB does not store title-of-class, so a *genuine* preferred whose
# name matches the common issuer will re-resolve to the common ticker here;
# full common/preferred separation comes from re-ingesting (class is now
# captured) and is applied at ingest by resolve_ticker's title_class branch.
_TICKER_FIX_CACHE = {}

def corrected_ticker(stored, cusip, name):
    if not _is_preferred_ticker(stored):
        return stored
    key = (stored, cusip or "", name or "")
    if key in _TICKER_FIX_CACHE:
        return _TICKER_FIX_CACHE[key]
    fixed = resolve_ticker(cusip, name)          # no class available from DB
    if _is_preferred_ticker(fixed) or fixed.startswith("~"):
        fixed = stored                            # nothing better — keep as-is
    _TICKER_FIX_CACHE[key] = fixed
    return fixed


def resolve_sector(ticker, name):
    """
    Return a sector string for a ticker.
    Strategy:
      1. Static overrides (highest priority)
      2. Dynamic ticker→sector cache (built from SEC name keywords)
      3. Name keyword fallback
      4. "Other" as last resort
    """
    if not _LOOKUP_LOADED:
        load_sec_ticker_file()

    # 1. Static overrides
    if ticker in TICKER_OVERRIDES:
        return TICKER_OVERRIDES[ticker]

    # 2. Yahoo Finance cache (persisted across runs, keyed by ticker OR name)
    cached = _YF_SECTORS.get(ticker.upper())
    if cached:
        if isinstance(cached, dict) and cached.get("sector") and cached["sector"] != "Other":
            return cached["sector"]
        elif isinstance(cached, str) and cached != "Other":
            return cached

    # 3. Dynamic SEC name-keyword cache
    if ticker in _TICKER_SECTOR:
        return _TICKER_SECTOR[ticker]

    # 4. Name keywords
    sector = _sector_from_name(name.lower())
    if sector:
        _TICKER_SECTOR[ticker] = sector
        return sector

    return "Other"


# Yahoo Finance sector map — persisted to disk, never expires
# (sectors rarely change; we only add new entries, never overwrite existing ones)
_YF_SECTOR_CACHE_PATH = None   # set after CACHE_DIR is known
_YF_SECTORS = {}               # ticker → sector, loaded from disk once
_YF_LOCK = threading.Lock()

# Map Yahoo Finance sector strings → our sector labels
YF_SECTOR_MAP = {
    "Technology":              "Technology",
    "Consumer Cyclical":       "Consumer Discr.",
    "Consumer Defensive":      "Consumer Staples",
    "Healthcare":              "Health Care",
    "Financial Services":      "Financials",
    "Energy":                  "Energy",
    "Industrials":             "Industrials",
    "Communication Services":  "Communication",
    "Basic Materials":         "Materials",
    "Real Estate":             "Real Estate",
    "Utilities":               "Utilities",
}


def _load_yf_cache():
    """Load the persisted Yahoo Finance sector cache from disk."""
    global _YF_SECTORS, _YF_SECTOR_CACHE_PATH
    _YF_SECTOR_CACHE_PATH = CACHE_DIR / "yf_sectors.json"
    if _YF_SECTOR_CACHE_PATH.exists():
        try:
            with open(_YF_SECTOR_CACHE_PATH) as f:
                _YF_SECTORS.update(json.load(f))
            print(f"  [YF] Loaded {len(_YF_SECTORS):,} cached Yahoo Finance sectors")
        except Exception:
            pass


def _save_yf_cache():
    """Persist Yahoo Finance sector cache to disk."""
    if _YF_SECTOR_CACHE_PATH:
        try:
            CACHE_DIR.mkdir(exist_ok=True)
            with open(_YF_SECTOR_CACHE_PATH, "w") as f:
                json.dump(_YF_SECTORS, f)
        except Exception:
            pass


def _yf_search(company_name):
    """
    Search Yahoo Finance by company name.
    Returns (ticker, sector) or (None, None) on failure.

    Uses the /v1/finance/search endpoint which returns ticker, quoteType,
    sector and industry directly in the search result — only ONE HTTP call needed.
    No second quoteSummary call required.
    """
    import urllib.parse
    query = urllib.parse.quote(company_name.strip())
    # Use the quote endpoint instead of search — it returns sector inline
    url = (
        f"https://query1.finance.yahoo.com/v1/finance/search"
        f"?q={query}&quotesCount=3&newsCount=0&enableFuzzyQuery=false"
        f"&quotesQueryId=tss_match_phrase_query"
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://finance.yahoo.com/",
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())

        quotes = data.get("quotes", [])
        for q in quotes[:5]:
            qt = q.get("quoteType", "")
            if qt not in ("EQUITY", "ETF", "MUTUALFUND"):
                continue

            ticker = q.get("symbol", "")
            if not ticker:
                continue

            # ETFs always → "ETF"
            if qt == "ETF":
                return ticker, "ETF"

            # Try to get sector from the search result directly
            # Yahoo sometimes includes 'sector' in search results
            yf_sector = q.get("sector", "") or q.get("industry", "")
            sector = YF_SECTOR_MAP.get(yf_sector)

            # If not in search result, try the quote summary (single extra call)
            if not sector:
                sector = _yf_sector_from_ticker(ticker)

            return ticker, sector

    except Exception:
        pass
    return None, None


def _yf_sector_from_ticker(ticker):
    """
    Fetch sector for a known ticker via Yahoo Finance.
    Uses the v6/finance/quoteSummary endpoint which is more reliable
    and returns sector without requiring a cookie/crumb.
    Returns our sector label or None.
    """
    # Try v6 first (more reliable, no crumb needed)
    for url in [
        f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=assetProfile",
        f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=assetProfile",
    ]:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "application/json",
            "Referer": "https://finance.yahoo.com/",
        }
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as r:
                data = json.loads(r.read())
            result = data.get("quoteSummary", {}).get("result")
            if result:
                profile = result[0].get("assetProfile", {})
                yf_sector = profile.get("sector", "")
                mapped = YF_SECTOR_MAP.get(yf_sector)
                if mapped:
                    return mapped
        except Exception:
            continue
    return None


def enrich_sectors_via_yahoo(holdings):
    """
    For every holding still labeled 'Other' (or with a derived/guessed ticker),
    search Yahoo Finance by the company NAME from the SEC filing.
    This resolves both the correct ticker AND the sector in one shot —
    no static CUSIP maps needed for this path.

    Results cached permanently in cache/yf_sectors.json so each company
    is only ever looked up once.
    """
    global _YF_SECTORS

    if _YF_SECTOR_CACHE_PATH is None:
        _load_yf_cache()

    # Build lookup list: holdings whose sector is still Other
    # Key by the SEC company name (not the derived ticker) for accuracy
    need_lookup = []
    for h in holdings:
        name = h.get("name", "")
        # Cache key is the normalized company name
        cache_key = name.strip().upper()
        if (h.get("sector") == "Other" or h.get("ticker", "").startswith("~")) and cache_key not in _YF_SECTORS:
            need_lookup.append(h)
        elif cache_key in _YF_SECTORS:
            cached = _YF_SECTORS[cache_key]
            if isinstance(cached, dict):
                # Stored as {ticker, sector} — always apply ticker if we have one
                if cached.get("ticker") and (h.get("ticker","").startswith("~") or not h.get("ticker")):
                    h["ticker"] = cached["ticker"]
                if cached.get("sector") and cached["sector"] != "Other":
                    h["sector"] = cached["sector"]
            elif isinstance(cached, str) and cached != "Other":
                h["sector"] = cached

    if not need_lookup:
        return

    print(f"  [YF] Searching {len(need_lookup)} companies by name via Yahoo Finance…")

    def lookup_one(h):
        name = h.get("name", "")
        cache_key = name.strip().upper()

        ticker, sector = _yf_search(name)

        result = {
            "ticker": ticker or h.get("ticker", ""),
            "sector": sector or "Other"
        }

        with _YF_LOCK:
            _YF_SECTORS[cache_key] = result
            # Also cache by ticker for future resolve_sector calls
            if ticker:
                _YF_SECTORS[ticker.upper()] = result

        if ticker:
            h["ticker"] = ticker
        if sector and sector != "Other":
            h["sector"] = sector

        time.sleep(0.08)  # polite throttle

    # Run in parallel batches of 20
    threads = []
    cap = min(len(need_lookup), 300)  # never more than 300 per fund
    for h in need_lookup[:cap]:
        t = threading.Thread(target=lookup_one, args=(h,))
        t.daemon = True
        t.start()
        threads.append(t)
        if len(threads) >= 20:
            for th in threads:
                th.join()
            threads = []
    for t in threads:
        t.join()

    _save_yf_cache()
    resolved = sum(1 for h in holdings if h.get("sector", "Other") != "Other")
    print(f"  [YF] Done — {resolved}/{len(holdings)} holdings have a sector")


def enrich_with_tickers(holdings):
    """
    Map CUSIPs → tickers, assign sectors, then fill gaps via Yahoo Finance.
    """
    if not _LOOKUP_LOADED:
        load_sec_ticker_file()

    # Load YF cache early so resolve_sector can use it
    if _YF_SECTOR_CACHE_PATH is None:
        _load_yf_cache()

    total_value = sum(h["value"] for h in holdings) or 1
    result = []

    for h in holdings:
        ticker = resolve_ticker(h["cusip"], h["name"], h.get("class", ""))
        h["ticker"] = ticker
        h["pct"]    = round(h["value"] / total_value * 100, 2)
        result.append(h)

    # Merge duplicate tickers (sub-manager splits).
    # Puts/calls get a unique key so they stay separate from stock positions
    # but are included in total AUM calculation.
    seen = {}
    for h in result:
        t = h["ticker"]
        pc = h.get("putCall")
        if t.startswith("~"):
            key = t + "_" + h.get("cusip","")[:6] + ("_" + pc if pc else "")
            seen[key] = dict(h)
        elif pc:
            key = t + "_" + pc
            if key in seen:
                seen[key]["value"]  += h["value"]
                seen[key]["shares"] += h["shares"]
            else:
                seen[key] = dict(h)
        elif t in seen and not seen[t].get("putCall"):
            seen[t]["value"]  += h["value"]
            seen[t]["shares"] += h["shares"]
        else:
            seen[t] = dict(h)

    merged = list(seen.values())
    total_value = sum(h["value"] for h in merged) or 1
    for h in merged:
        h["pct"] = round(h["value"] / total_value * 100, 2)

    stocks  = sorted([h for h in merged if not h.get("putCall")], key=lambda x: x["pct"], reverse=True)
    options = sorted([h for h in merged if h.get("putCall")],     key=lambda x: x["pct"], reverse=True)
    return stocks + options


def guess_sector(ticker, name):
    """Public alias kept for compatibility — delegates to resolve_sector."""
    return resolve_sector(ticker, name)


# ─────────────────────────────────────────────
#  CACHE HELPERS
# ─────────────────────────────────────────────

def cache_path(investor_id, which):
    return CACHE_DIR / f"{investor_id}_{which}.json"


def cache_valid(path):
    """Cache is valid only if it is less than 24h old AND from today's run date."""
    if not path.exists():
        return False

    if (time.time() - path.stat().st_mtime) >= CACHE_TTL:
        return False

    # Force a fresh SEC check on a new calendar day, even if the cache is
    # technically less than 24 hours old. This matches the app behavior you want:
    # every day, use today's date to find the closest/latest available 13F.
    try:
        cached = load_cache(path)
        if cached.get("cacheRunDate") != date.today().isoformat():
            return False
    except Exception:
        return False

    return True


def load_cache(path):
    with open(path) as f:
        return json.load(f)


def save_cache(path, data):
    CACHE_DIR.mkdir(exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f)


# ─────────────────────────────────────────────
#  PRICE PERFORMANCE SINCE FILING
#  Uses yfinance to get price at filing date
#  and current price. Cached to disk.
# ─────────────────────────────────────────────

_PRICE_CACHE: dict = {}          # in-memory session cache  (ticker:date → price/dict)
_PRICE_CACHE_LOADED = False
_PRICE_CACHE_DIRTY: set = set()  # keys that need flushing to SQLite
_PRICE_DB_PATH = None            # set on first load

def _init_price_db():
    """Create the price_cache SQLite table (and migrate the old JSON file if present)."""
    global _PRICE_DB_PATH
    CACHE_DIR.mkdir(exist_ok=True)
    _PRICE_DB_PATH = CACHE_DIR / "price_cache.db"
    conn = sqlite3.connect(str(_PRICE_DB_PATH))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_cache (
            cache_key  TEXT PRIMARY KEY,
            value      TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()
    # One-time migration from the old price_cache.json
    json_path = CACHE_DIR / "price_cache.json"
    if json_path.exists():
        try:
            with open(json_path) as f:
                old = json.load(f)
            if old:
                conn.executemany(
                    "INSERT OR IGNORE INTO price_cache (cache_key, value) VALUES (?, ?)",
                    [(k, json.dumps(v)) for k, v in old.items()]
                )
                conn.commit()
                print(f"  [CACHE] Migrated {len(old)} prices from JSON → SQLite")
            json_path.rename(json_path.with_suffix(".json.migrated"))
        except Exception as e:
            print(f"  [CACHE] Warning: JSON migration failed: {e}")
    conn.close()

def _load_price_cache():
    global _PRICE_CACHE, _PRICE_CACHE_LOADED
    _PRICE_CACHE_LOADED = True
    _init_price_db()
    try:
        conn = sqlite3.connect(str(_PRICE_DB_PATH))
        rows = conn.execute("SELECT cache_key, value FROM price_cache").fetchall()
        conn.close()
        for k, v in rows:
            try:
                _PRICE_CACHE[k] = json.loads(v) if v is not None else None
            except (json.JSONDecodeError, TypeError):
                _PRICE_CACHE[k] = None
        print(f"  [CACHE] Loaded {len(_PRICE_CACHE):,} cached prices from SQLite")
    except Exception as e:
        print(f"  [CACHE] Warning: could not load price cache: {e}")

def _save_price_cache():
    """Flush only the NEW entries written this session to SQLite."""
    if not _PRICE_CACHE_DIRTY or not _PRICE_DB_PATH:
        return
    try:
        conn = sqlite3.connect(str(_PRICE_DB_PATH))
        conn.executemany(
            "INSERT OR REPLACE INTO price_cache (cache_key, value) VALUES (?, ?)",
            [(k, json.dumps(_PRICE_CACHE.get(k))) for k in _PRICE_CACHE_DIRTY
             if k in _PRICE_CACHE]
        )
        conn.commit()
        conn.close()
        n = len(_PRICE_CACHE_DIRTY)
        _PRICE_CACHE_DIRTY.clear()
        print(f"  [CACHE] Saved {n} new price entries to SQLite")
    except Exception as e:
        print(f"  [CACHE] Warning: could not save price cache: {e}")

def _get_price_on_date(ticker, date_str):
    """Get closing price for ticker on or just after date_str. Cached."""
    if not _YF_AVAILABLE:
        return None
    if not _PRICE_CACHE_LOADED:
        _load_price_cache()
    key = f"{ticker}:{date_str}"
    if key in _PRICE_CACHE:
        return _PRICE_CACHE[key]
    try:
        start = date.fromisoformat(date_str)
        end   = start + timedelta(days=7)
        hist  = yf.Ticker(ticker).history(start=start.isoformat(), end=end.isoformat(), auto_adjust=True)
        if hist.empty:
            _PRICE_CACHE[key] = None; _PRICE_CACHE_DIRTY.add(key)
            return None
        price = round(float(hist["Close"].iloc[0]), 4)
        _PRICE_CACHE[key] = price; _PRICE_CACHE_DIRTY.add(key)
        return price
    except Exception as e:
        print(f"  [PERF] Price lookup error {ticker} @ {date_str}: {e}")
        _PRICE_CACHE[key] = None; _PRICE_CACHE_DIRTY.add(key)
        return None

def _get_current_price(ticker):
    """Get latest closing price for ticker. Cached with today's date as key."""
    if not _YF_AVAILABLE:
        return None
    if not _PRICE_CACHE_LOADED:
        _load_price_cache()
    today_key = f"{ticker}:today:{date.today().isoformat()}"
    if today_key in _PRICE_CACHE:
        return _PRICE_CACHE[today_key]
    try:
        hist = yf.Ticker(ticker).history(period="2d", auto_adjust=True)
        if hist.empty:
            _PRICE_CACHE[today_key] = None; _PRICE_CACHE_DIRTY.add(today_key)
            return None
        price = round(float(hist["Close"].iloc[-1]), 4)
        _PRICE_CACHE[today_key] = price; _PRICE_CACHE_DIRTY.add(today_key)
        return price
    except Exception as e:
        print(f"  [PERF] Current price error {ticker}: {e}")
        _PRICE_CACHE[today_key] = None; _PRICE_CACHE_DIRTY.add(today_key)
        return None

# ─────────────────────────────────────────────
#  13F-IMPLIED PRICE (FALLBACK ONLY)
#  When yfinance has no price for a name on a backtest boundary
#  (de-listed, mis-mapped/re-used ticker, foreign line, CUSIP-only
#  placeholder), we derive the price the managers themselves implied:
#      per-share price = value / shares
#  `value` is stored in $millions and `shares` is a raw count, so the
#  per-share figure is value * 1e6 / shares. Every investor who held the
#  name that quarter gives one estimate; we take the MEDIAN across all of
#  them, so a single mis-detected unit or bad row can't move the result.
#  A sanity band drops absurd values before the median.
#
#  NOTE: this is the quarter-END mark (the 13F report date), not the
#  filing-date close, so it is a proxy used ONLY where no market price
#  exists. Where both boundaries of a name fall back to implied, the
#  return is a consistent quarter-end -> quarter-end comparison.
# ─────────────────────────────────────────────
_IMPLIED_CACHE = {}

# A 13F-implied mark is a RAW (non-split-adjusted) value/shares figure, so it
# cannot bridge a stock split, and a mis-mapped ticker (e.g. "C-PR" catching
# different Citigroup securities across quarters) can make value/shares jump
# between quarters for reasons that are not real returns. yfinance market
# prices are split-adjusted and don't have this problem. So when a name's
# return relies on an implied mark on EITHER leg, we reject moves outside a
# plausible single-quarter band as artifacts (Citi's 2011 1-for-10 reverse
# split alone would fake a +900% leg) and drop the name that period.
IMPLIED_RET_MAX =  1.5     # +150% in one quarter via an implied mark -> artifact
IMPLIED_RET_MIN = -0.75    # -75%  in one quarter via an implied mark -> artifact

def _implied_artifact(src0, src1, ret):
    """True if `ret` came from an implied mark and is implausibly large."""
    if src0 != "implied" and src1 != "implied":
        return False
    return ret > IMPLIED_RET_MAX or ret < IMPLIED_RET_MIN

def _implied_price(ticker, quarter):
    """Median 13F-implied price (value*1e6/shares) for `ticker` in `quarter`.

    Aggregates every SH holding of the ticker across all investors that
    quarter. Options (put_call set) and zero/negative rows are excluded.
    Returns a float, or None if the quarter has no usable share holding.
    """
    if not ticker or not quarter:
        return None
    key = (ticker.upper(), quarter)
    if key in _IMPLIED_CACHE:
        return _IMPLIED_CACHE[key]

    price = None
    if _db_available():
        try:
            conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
            rows = conn.execute("""
                SELECT h.value, h.shares
                FROM   holdings h
                JOIN   filings  f ON h.filing_id = f.id
                WHERE  UPPER(h.ticker) = ?
                  AND  f.quarter = ?
                  AND  (h.put_call IS NULL OR h.put_call = '')
                  AND  h.shares > 0
                  AND  h.value  > 0
            """, (ticker.upper(), quarter)).fetchall()
            conn.close()

            implied = []
            for value_m, shares in rows:
                try:
                    p = (float(value_m) * 1_000_000.0) / float(shares)
                except (TypeError, ValueError, ZeroDivisionError):
                    continue
                if 0.01 <= p <= 1_000_000.0:   # sanity band: drop unit glitches
                    implied.append(p)

            if implied:
                implied.sort()
                n = len(implied)
                price = (implied[n // 2] if n % 2
                         else (implied[n // 2 - 1] + implied[n // 2]) / 2.0)
                price = round(price, 4)
        except Exception as e:
            print(f"  [PERF] Implied-price error {ticker} {quarter}: {e}")
            price = None

    _IMPLIED_CACHE[key] = price
    return price


def _price_for(ticker, date_str, quarter):
    """Market price (yfinance) if available, else 13F-implied fallback.

    Returns (price, source) where source is 'market', 'implied', or None.
    Market is always tried first; the implied fallback fires ONLY when the
    market price is missing — i.e. only for names you don't have a price for.
    """
    if not ticker.startswith("~"):            # "~" = CUSIP-only, never on yfinance
        p = _get_price_on_date(ticker, date_str)
        if p and p > 0:
            return p, "market"
    p = _implied_price(ticker, quarter)
    if p and p > 0:
        return p, "implied"
    return None, None

# ─────────────────────────────────────────────
#  CONSENSUS BACKTEST ENGINE
#  Equal-weight, quarterly-rebalanced backtest of the top-N consensus
#  baskets. Prices come from yfinance via _get_price_on_date(), which
#  returns the SPLIT/DIVIDEND-ADJUSTED close — so splits and share-class
#  artefacts that made value/shares unreliable do not distort returns.
#
#  Timing model (documented for the UI so the math is auditable):
#    - The basket for quarter Q holds the stocks that were in the top-N
#      consensus for that reported quarter.
#    - To avoid quarter-end look-ahead, the basket starts on the latest
#      filing date among the selected investors who filed for quarter Q
#      (the frontend sends this as entryDate).
#    - We hold that basket until the latest filing date for the NEXT quarter,
#      then rebalance to the new top-N (selling names that dropped out,
#      buying any that entered, re-equal-weighting).
#    - Each name's period return is adjClose(next quarter filing date) /
#      adjClose(current quarter filing date) − 1. The basket return is the
#      simple average (equal weight) across the names we could actually price.
#    - Names with no usable price (de-listed, foreign line, CUSIP-only
#      "~" placeholders) are dropped from THAT period and reported, so the
#      coverage is always visible rather than silently fabricated.
# ─────────────────────────────────────────────

def _quarter_end_date(label):
    """'Q1 2025' -> '2025-03-31'. Returns None if unparseable."""
    try:
        q, y = str(label).strip().split()
        y = int(y)
        ends = {"Q1": (3, 31), "Q2": (6, 30), "Q3": (9, 30), "Q4": (12, 31)}
        m, d = ends[q.upper()]
        return date(y, m, d).isoformat()
    except Exception:
        return None


def _valid_iso_date(value):
    """Return YYYY-MM-DD if value is a valid ISO date, else None."""
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date().isoformat()
    except Exception:
        return None


def run_consensus_backtest(baskets, include_partial=False):
    """
    baskets: list of {
        "quarter": "Q1 2025",
        "tickers": ["MSFT", ...],
        "entryDate": "2025-05-15"   # latest filing date for selected investors
    } (any order).
    Returns a dict the frontend renders as a chart + an auditable breakdown.
    """
    if not _YF_AVAILABLE:
        return {"ok": False,
                "error": "yfinance is not installed on the server. Run: pip install yfinance"}

    # Normalise + sort oldest -> newest by investable entry date.
    # entryDate should be the latest filing date among the selected investors
    # for the reported quarter. If older frontends do not send it, fall back
    # to quarter-end so the endpoint remains backward-compatible.
    norm = []
    for b in baskets or []:
        qe = _quarter_end_date(b.get("quarter", ""))
        if not qe:
            continue
        entry = (_valid_iso_date(b.get("entryDate")) or
                 _valid_iso_date(b.get("filingDate")) or
                 qe)
        tickers = []
        for t in (b.get("tickers") or []):
            t = str(t or "").strip().upper()
            if t and t not in ("—", "-"):
                tickers.append(t)
        investors = [str(n) for n in (b.get("investors") or []) if n]
        norm.append({"quarter": b.get("quarter"), "qend": qe, "entry": entry,
                     "tickers": tickers, "investors": investors})
    norm.sort(key=lambda x: x["entry"])

    if len(norm) < 2:
        return {"ok": False,
                "error": "Need at least two quarters with holdings to compute a return."}

    periods = []
    cumulative = 1.0

    def good_price(x):
        try:
            x = float(x)
            return math.isfinite(x) and x > 0
        except Exception:
            return False

    def price_basket(tickers, start, end, q_start, q_end):
        """Return (constituents, dropped, dropped_implausible) for start->end.

        Each leg is priced market-first; the 13F-implied price fills in only
        where yfinance has nothing (q_start / q_end give the fallback quarter).
        Implied-derived returns outside a plausible band (split / identity
        artifacts) are rejected and reported separately, never averaged in.
        """
        constituents, dropped, dropped_implausible = [], [], []
        for t in tickers:
            p0, s0 = _price_for(t, start, q_start)
            p1, s1 = _price_for(t, end,   q_end)
            if good_price(p0) and good_price(p1):
                p0 = float(p0)
                p1 = float(p1)
                ret = p1 / p0 - 1.0

                if not math.isfinite(ret):
                    dropped.append(t)
                    continue
                if _implied_artifact(s0, s1, ret):
                    dropped_implausible.append({"ticker": t, "p0": round(p0, 4),
                                                "p1": round(p1, 4),
                                                "ret": round(ret, 6)})
                    continue
                constituents.append({"ticker": t, "p0": p0, "p1": p1,
                                     "ret": round(ret, 6),
                                     "src0": s0, "src1": s1,
                                     "implied": (s0 == "implied" or s1 == "implied")})
            else:
                dropped.append(t)
        return constituents, dropped, dropped_implausible

    for i in range(len(norm) - 1):
        cur, nxt = norm[i], norm[i + 1]
        constituents, dropped, dropped_implausible = price_basket(
            cur["tickers"], cur["entry"], nxt["entry"], cur["quarter"], nxt["quarter"])
        basket_ret = (sum(c["ret"] for c in constituents) / len(constituents)
                      if constituents else 0.0)
        cumulative *= (1.0 + basket_ret)
        periods.append({
            "from_quarter": cur["quarter"], "to_quarter": nxt["quarter"],
            "from": cur["entry"], "to": nxt["entry"],
            "from_quarter_end": cur["qend"], "to_quarter_end": nxt["qend"],
            "investors": cur["investors"],
            "n_in": len(cur["tickers"]), "n_priced": len(constituents),
            "n_implied": sum(1 for c in constituents if c["implied"]),
            "dropped": dropped,
            "dropped_implausible": dropped_implausible,
            "constituents": sorted(constituents, key=lambda c: -c["ret"]),
            "basket_ret": round(basket_ret, 6),
            "cumulative": round(cumulative, 6),
            "partial": False,
        })

    # Optional: price the most-recent basket forward to today (incomplete quarter).
    if include_partial:
        last = norm[-1]
        today = date.today().isoformat()
        if today > last["entry"]:
            constituents, dropped, dropped_implausible = [], [], []
            for t in last["tickers"]:
                p0, s0 = _price_for(t, last["entry"], last["quarter"])
                p1 = _get_current_price(t) if not t.startswith("~") else None
                if good_price(p0) and good_price(p1):
                    p0 = float(p0)
                    p1 = float(p1)
                    ret = p1 / p0 - 1.0

                    if not math.isfinite(ret):
                        dropped.append(t)
                        continue
                    if _implied_artifact(s0, "market", ret):
                        dropped_implausible.append({"ticker": t, "p0": round(p0, 4),
                                                    "p1": round(p1, 4),
                                                    "ret": round(ret, 6)})
                        continue
                    constituents.append({"ticker": t, "p0": p0, "p1": p1,
                                         "ret": round(ret, 6),
                                         "src0": s0, "src1": "market",
                                         "implied": (s0 == "implied")})
                else:
                    dropped.append(t)
            basket_ret = (sum(c["ret"] for c in constituents) / len(constituents)
                          if constituents else 0.0)
            cumulative *= (1.0 + basket_ret)
            periods.append({
                "from_quarter": last["quarter"], "to_quarter": "now",
                "from": last["entry"], "to": today,
                "from_quarter_end": last["qend"], "to_quarter_end": None,
                "investors": last["investors"],
                "n_in": len(last["tickers"]), "n_priced": len(constituents),
                "n_implied": sum(1 for c in constituents if c["implied"]),
                "dropped": dropped,
                "dropped_implausible": dropped_implausible,
                "constituents": sorted(constituents, key=lambda c: -c["ret"]),
                "basket_ret": round(basket_ret, 6),
                "cumulative": round(cumulative, 6),
                "partial": True,
            })

    _save_price_cache()

    avg_cov = (sum(p["n_priced"] for p in periods) /
               max(1, sum(p["n_in"] for p in periods)))
    return {
        "ok": True,
        "start_date": norm[0]["entry"],
        "end_date": periods[-1]["to"] if periods else norm[-1]["entry"],
        "n_periods": len(periods),
        "coverage": round(avg_cov, 4),
        "n_implied": sum(p.get("n_implied", 0) for p in periods),
        "cumulative_return": round(cumulative - 1.0, 6),
        "method": ("Equal-weight; rebalanced every quarter to that quarter's "
                   "top-15 consensus; entries use the latest 13F filing date "
                   "among the selected investors for the reported quarter; "
                   "held filing-date to next filing-date; split/dividend-adjusted "
                   "closes. Names with no market price fall back to the median "
                   "13F-implied price (value/shares across all holders that "
                   "quarter) — a quarter-end mark used only where no market "
                   "price exists. Names that can't be priced either way are "
                   "dropped per-period and the survivors re-equal-weighted."),
        "periods": periods,
    }


# ─────────────────────────────────────────────
#  CAGR OPTIMIZER
#  Searches investor subsets × top-N to maximise CAGR. The full space
#  (2^N subsets × top-N values) is astronomically large AND maximising
#  in-sample CAGR overfits, so this does a BOUNDED search (greedy / random)
#  and reports an OUT-OF-SAMPLE CAGR on a held-out tail of history, so the
#  user can judge whether a "winner" is real or just data-mined.
#
#  For speed and fair comparison, every candidate is priced on the SAME
#  per-quarter entry date (the latest filing date across the WHOLE pool),
#  so the search measures stock selection, not timing luck, and the price
#  cache is reused across thousands of evaluations. The chosen combination
#  is then re-run through the normal backtest (exact per-subset timing) when
#  the user clicks "Apply".
# ─────────────────────────────────────────────
import re as _re_opt

_OPT_CLASS_RE  = _re_opt.compile(r'\b(CL(ASS)?|SER(IES)?|COM|COMMON|ADR|ADS|SPON(SORED)?|WT|WTS|RT|RTS|UNIT|UNITS|PFD|PREFERRED)\b.*$')
_OPT_SUFFIX_RE = _re_opt.compile(r'\b(INC|CORP|CORPORATION|CO|COS|COMPANIES|COMPANY|LTD|LIMITED|LLC|LP|PLC|NV|SA|AG|HLDG|HLDGS|HOLDING|HOLDINGS|GROUP|GRP|TR|TRUST|THE|NEW|REIT|INTL|INTERNATIONAL)\b')

def _issuer_key(name):
    """Collapse share classes to a single issuer — mirrors the frontend."""
    s = (name or "").upper()
    s = _OPT_CLASS_RE.sub(" ", s)
    s = _re_opt.sub(r'[.,&/()]', ' ', s)
    s = _OPT_SUFFIX_RE.sub(" ", s)
    s = _re_opt.sub(r'\s+', ' ', s).strip()
    return s or (name or "").upper().strip()

def _opt_load_pool(ids):
    """{inv_id: {quarter: {"date":filing_date, "holds":[{name,ticker,value}]}}}.
    Keeps only the latest filing per quarter; drops options and empty rows;
    applies the read-time ticker correction so scoring matches the backtest."""
    pool = {}
    if not _db_available():
        return pool
    conn = _db_connect()
    try:
        for inv_id in ids:
            filings = conn.execute(
                "SELECT id, filing_date, quarter FROM filings WHERE investor_id=? ORDER BY filing_date DESC",
                (inv_id,)).fetchall()
            qmap = {}
            for f in filings:
                if f["quarter"] in qmap:
                    continue
                rows = conn.execute(
                    "SELECT name, cusip, ticker, value, put_call FROM holdings WHERE filing_id=? ORDER BY pct DESC",
                    (f["id"],)).fetchall()
                holds = []
                for r in rows:
                    if r["put_call"] or not r["name"] or (r["value"] or 0) <= 0:
                        continue
                    holds.append({"name": r["name"],
                                  "ticker": corrected_ticker(r["ticker"], r["cusip"], r["name"]),
                                  "value": r["value"] or 0.0})
                qmap[f["quarter"]] = {"date": f["filing_date"], "holds": holds}
            if qmap:
                pool[inv_id] = qmap
    finally:
        conn.close()
    return pool

def _opt_pool_entry_dates(pool, ids):
    """Latest filing date per quarter across the whole pool (stable pricing)."""
    qd = {}
    for inv_id in ids:
        for q, d in pool.get(inv_id, {}).items():
            if d["date"] and (q not in qd or d["date"] > qd[q]):
                qd[q] = d["date"]
    return qd

def _opt_baskets(pool, ids, topn, entry_dates):
    """Consensus baskets for a subset — mirrors the frontend consensus exactly."""
    min_holders = 1 if len(ids) <= 1 else 2
    quarters = sorted({q for i in ids for q in pool.get(i, {})},
                      key=lambda q: entry_dates.get(q, ""))
    baskets = []
    for q in quarters:
        by = {}
        for inv_id in ids:
            qd = pool.get(inv_id, {}).get(q)
            if not qd:
                continue
            for h in qd["holds"][:topn]:
                key = _issuer_key(h["name"])
                e = by.setdefault(key, {"tc": {}, "holders": {}})
                t = h["ticker"]
                if t and not t.startswith("~"):
                    e["tc"][t] = e["tc"].get(t, 0) + 1
                e["holders"][inv_id] = e["holders"].get(inv_id, 0.0) + h["value"]
        ranked = []
        for e in by.values():
            if len(e["holders"]) < min_holders:
                continue
            ticker = max(e["tc"].items(), key=lambda kv: kv[1])[0] if e["tc"] else "—"
            ranked.append((len(e["holders"]), sum(e["holders"].values()), ticker))
        ranked.sort(key=lambda r: (-r[0], -r[1]))
        tickers = [t for (_c, _v, t) in ranked[:15] if t and t not in ("—", "-")]
        if tickers:
            baskets.append({"quarter": q, "entryDate": entry_dates.get(q),
                            "tickers": tickers, "investors": []})
    return baskets

def _opt_cagr(res):
    if not res or not res.get("ok"):
        return None
    try:
        d0 = date.fromisoformat(res["start_date"]); d1 = date.fromisoformat(res["end_date"])
        days = (d1 - d0).days
        if days <= 0:
            return None
        return (1.0 + res["cumulative_return"]) ** (365.25 / days) - 1.0
    except Exception:
        return None

def _opt_maxdd(res):
    peak, mdd = 1.0, 0.0
    for p in (res.get("periods") or []):
        c = p.get("cumulative", 1.0)
        peak = max(peak, c)
        if peak > 0:
            mdd = min(mdd, c / peak - 1.0)
    return round(mdd, 4)

def _opt_evaluate(pool, ids, topn, entry_dates, split_idx, include_partial):
    """Score one (subset, topn) on train (and held-out test if split set)."""
    baskets = _opt_baskets(pool, ids, topn, entry_dates)
    if len(baskets) < 2:
        return None
    train_b = baskets[:split_idx] if (1 < split_idx < len(baskets)) else baskets
    test_b  = baskets[split_idx:] if (1 < split_idx < len(baskets)) else []
    train = run_consensus_backtest(train_b, include_partial=False)
    train_cagr = _opt_cagr(train)
    if train_cagr is None:
        return None
    m = {"topn": topn, "n_investors": len(ids), "investor_ids": list(ids),
         "train_cagr": round(train_cagr, 6),
         "train_return": round(train.get("cumulative_return", 0.0), 6),
         "quarters": train.get("n_periods", 0),
         "coverage": train.get("coverage", 0.0),
         "n_implied": train.get("n_implied", 0),
         "max_drawdown": _opt_maxdd(train),
         "test_cagr": None, "test_return": None, "test_quarters": 0}
    if len(test_b) >= 2:
        test = run_consensus_backtest(test_b, include_partial=include_partial)
        tc = _opt_cagr(test)
        if tc is not None:
            m["test_cagr"]    = round(tc, 6)
            m["test_return"]  = round(test.get("cumulative_return", 0.0), 6)
            m["test_quarters"] = test.get("n_periods", 0)
    return m

def run_cagr_optimizer(req):
    ids_all  = [str(x) for x in (req.get("investorIds") or []) if x]
    id_name  = {str(k): v for k, v in (req.get("investorNames") or {}).items()}
    min_inv  = max(1, int(req.get("minInvestors", 1)))
    topn_min = max(1, int(req.get("topnMin", 10)))
    topn_max = max(topn_min, int(req.get("topnMax", topn_min)))
    topn_step = max(1, int(req.get("topnStep", max(1, (topn_max - topn_min) // 4 or 1))))
    method   = (req.get("method") or "greedy").lower()
    budget   = max(20, min(int(req.get("budget", 800)), 6000))
    oos_frac = float(req.get("oosFraction", 0.3))
    include_partial = bool(req.get("includePartial", False))

    pool = _opt_load_pool(ids_all)
    ids_all = [i for i in ids_all if i in pool]
    if len(ids_all) < min_inv:
        return {"ok": False, "error":
                f"Only {len(ids_all)} of the selected investors have filing data; "
                f"need at least the minimum of {min_inv}."}

    entry_dates = _opt_pool_entry_dates(pool, ids_all)
    topns = list(range(topn_min, topn_max + 1, topn_step))
    if topns[-1] != topn_max:
        topns.append(topn_max)

    ref_n = len(_opt_baskets(pool, ids_all, topns[0], entry_dates))
    split_idx = int(round(ref_n * (1.0 - oos_frac))) if (0.0 < oos_frac < 0.9 and ref_n >= 6) else 0

    evals, seen, results = 0, set(), []

    def consider(ids, topn):
        nonlocal evals
        key = (topn, tuple(sorted(ids)))
        if key in seen:
            return None
        seen.add(key)
        m = _opt_evaluate(pool, ids, topn, entry_dates, split_idx, include_partial)
        evals += 1
        if m:
            m["investors"] = [id_name.get(i, i) for i in ids]
            results.append(m)
        return m

    if method == "random":
        import random
        rng = random.Random(int(req.get("seed", 42)))
        sizes = list(range(min_inv, len(ids_all) + 1))
        attempts, max_attempts = 0, budget * 6 + 100
        while evals < budget and attempts < max_attempts:
            attempts += 1
            consider(rng.sample(ids_all, rng.choice(sizes)), rng.choice(topns))
    else:  # greedy forward selection per top-N
        for topn in topns:
            selected, remaining, best_score = [], list(ids_all), -1e18
            while remaining and evals < budget:
                best_cand, best_cand_score = None, -1e18
                for cand in list(remaining):
                    if evals >= budget:
                        break
                    m = consider(selected + [cand], topn)
                    sc = m["train_cagr"] if m else -1e18
                    if sc > best_cand_score:
                        best_cand_score, best_cand = sc, cand
                if best_cand is None:
                    break
                if len(selected) < min_inv or best_cand_score > best_score:
                    selected.append(best_cand); remaining.remove(best_cand)
                    best_score = max(best_score, best_cand_score)
                else:
                    break

    valid = [r for r in results if r["n_investors"] >= min_inv]
    valid.sort(key=lambda r: r["train_cagr"], reverse=True)

    return {
        "ok": True,
        "method": method,
        "evaluations": evals,
        "pool_size": len(ids_all),
        "topn_values": topns,
        "min_investors": min_inv,
        "oos_fraction": oos_frac if split_idx else 0.0,
        "train_quarters": split_idx if split_idx else ref_n,
        "test_quarters": (ref_n - split_idx) if split_idx else 0,
        "results": valid[:12],
        "note": ("Bounded search — NOT exhaustive. Ranked by IN-SAMPLE (train) CAGR. "
                 "A combination that scores far better in-sample than out-of-sample "
                 "is overfit to history and unlikely to repeat. 'Apply' re-runs it "
                 "through the full backtest with exact per-quarter filing-date timing."),
    }


def _get_fundamentals(ticker):
    """
    Fetch key fundamentals for a ticker via yfinance.
    Returns dict with marketCap, week52High, week52Low, targetPrice,
    forwardPE, trailingPE, evToEbitda, revenueGrowth.
    Cached in price_cache.json with today's date key (refreshes daily).
    """
    if not _YF_AVAILABLE:
        return {}
    if not _PRICE_CACHE_LOADED:
        _load_price_cache()
    key = f"{ticker}:fundamentals:{date.today().isoformat()}"
    if key in _PRICE_CACHE:
        return _PRICE_CACHE[key] or {}
    try:
        info = yf.Ticker(ticker).info
        result = {
            "marketCap":     info.get("marketCap"),
            "week52High":    info.get("fiftyTwoWeekHigh"),
            "week52Low":     info.get("fiftyTwoWeekLow"),
            "targetPrice":   info.get("targetMeanPrice"),
            "forwardPE":     info.get("forwardPE"),
            "trailingPE":    info.get("trailingPE"),
            "evToEbitda":    info.get("enterpriseToEbitda"),
            "revenueGrowth": info.get("revenueGrowth"),
        }
        _PRICE_CACHE[key] = result
        return result
    except Exception as e:
        print(f"  [FUND] Fundamentals error {ticker}: {e}")
        _PRICE_CACHE[key] = {}
        return {}


def enrich_performance(holdings, filing_date):
    """
    Add priceAtFiling, currentPrice, perfSinceFiling to each holding.
    Only processes top 50 positions to keep startup time reasonable.
    Returns weighted portfolio performance.
    """
    if not _YF_AVAILABLE or not filing_date:
        return None

    if not _PRICE_CACHE_LOADED:
        _load_price_cache()

    print(f"  [PERF] Fetching prices for top 50 holdings (filing: {filing_date})...")
    total_weight = 0.0
    weighted_perf = 0.0
    enriched = 0

    for h in holdings[:50]:
        ticker = h.get("ticker", "")
        if not ticker or ticker.startswith("~"):
            continue
        # Skip non-US tickers (contain dots like BRK.B is fine, but 931.DU etc. skip)
        if "." in ticker and not ticker.endswith(".B"):
            continue

        p_filing  = _get_price_on_date(ticker, filing_date)
        p_current = _get_current_price(ticker)

        if p_filing and p_current and p_filing > 0:
            underlying_perf = (p_current - p_filing) / p_filing * 100
            # Directional return of the *position*, not the underlying.
            # A PUT holder is bearish: they make money when the stock falls,
            # so the position's return is the inverse of the price change.
            # CALLs and share positions move with the underlying.
            if h.get("putCall") == "PUT":
                position_perf = -underlying_perf
            else:
                position_perf = underlying_perf
            perf = round(position_perf, 2)
            h["priceAtFiling"]  = p_filing
            h["currentPrice"]   = p_current
            h["perfSinceFiling"] = perf
            # Weight by portfolio % for overall return
            pct = h.get("pct", 0)
            weighted_perf += perf * pct / 100
            total_weight  += pct / 100
            enriched += 1
        else:
            h["priceAtFiling"]   = None
            h["currentPrice"]    = None
            h["perfSinceFiling"] = None

        # Fetch fundamentals for this ticker
        fund = _get_fundamentals(ticker)
        if fund:
            h["marketCap"]     = fund.get("marketCap")
            h["week52High"]    = fund.get("week52High")
            h["week52Low"]     = fund.get("week52Low")
            h["targetPrice"]   = fund.get("targetPrice")
            h["forwardPE"]     = fund.get("forwardPE")
            h["trailingPE"]    = fund.get("trailingPE")
            h["evToEbitda"]    = fund.get("evToEbitda")
            h["revenueGrowth"] = fund.get("revenueGrowth")
            # Compute % from 52w high
            if fund.get("week52High") and p_current and fund["week52High"] > 0:
                h["pctFrom52wHigh"] = round((p_current - fund["week52High"]) / fund["week52High"] * 100, 1)
            # Compute % upside to analyst target
            if fund.get("targetPrice") and p_current and p_current > 0:
                h["analystUpside"] = round((fund["targetPrice"] - p_current) / p_current * 100, 1)

    _save_price_cache()
    print(f"  [PERF] Enriched {enriched} positions")

    if total_weight > 0:
        portfolio_perf = round(weighted_perf / total_weight, 2)
        return portfolio_perf
    return None

import time

def background_enrich_all(batch_size=5, pause_sec=3):
    """Slowly enrich all investors in small batches AFTER startup.
    Runs one investor at a time; pauses between batches so memory from
    each yfinance fetch is reclaimed before the next batch starts.
    This is the safe alternative to enriching all 47 at once (which OOM'd)."""
    if not _YF_AVAILABLE:
        return
    ids = list(_all_data.keys())
    for i in range(0, len(ids), batch_size):
        batch = ids[i:i + batch_size]
        for inv_id in batch:
            with _data_lock:
                inv = _all_data.get(inv_id)
            if not inv or inv.get("_enriched"):
                continue
            try:
                perf = enrich_performance(inv.get("holdings", []), inv.get("filingDate", ""))
                with _data_lock:
                    inv["portfolioPerfSinceFiling"] = perf
                    inv["_enriched"] = True
                mem = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024
                print(f"  [BG-ENRICH] {inv_id} done · mem {mem:.0f} MB", flush=True)
            except Exception as e:
                print(f"  [BG-ENRICH] {inv_id} failed: {e}", flush=True)
            time.sleep(0.5)          # tiny gap between investors
        time.sleep(pause_sec)        # bigger pause between batches → lets memory settle


















# ─────────────────────────────────────────────
#  SQLITE DATABASE LAYER
#  Reads from 13f.db produced by ingest.py.
#  Falls back to live EDGAR fetch if DB is missing
#  (so the server still works before first ingest).
# ─────────────────────────────────────────────

import sqlite3

DB_PATH = Path(__file__).parent / "13f.db"


def _db_connect():
    """Open a read-only connection to the SQLite DB."""
    conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def _db_available():
    return DB_PATH.exists()


DB_DUMP_GZ  = Path(__file__).parent / "13f.db.gz"



def _rebuild_db_from_dump():
    """If 13f.db is missing but 13f.db.gz exists, stream-decompress it to disk.
    Streaming keeps peak memory near ~1 MB regardless of DB size (no OOM)."""
    if DB_PATH.exists():
        return
    if DB_DUMP_GZ.exists():
        print(f"\n  13f.db not found — decompressing {DB_DUMP_GZ.name}...")
        t0 = time.time()
        try:
            with gzip.open(str(DB_DUMP_GZ), "rb") as src, open(DB_PATH, "wb") as dst:
                shutil.copyfileobj(src, dst, length=1024 * 1024)  # 1 MB chunks
            size_mb = DB_PATH.stat().st_size / 1_000_000
            print(f"  Restored 13f.db ({size_mb:.0f} MB) in {time.time()-t0:.1f}s\n")
        except Exception as e:
            print(f"  ERROR decompressing DB: {e}")
            if DB_PATH.exists():
                DB_PATH.unlink()   # don't leave a half-written DB
    else:
        print("\n  13f.db not found (and no 13f.db.gz to rebuild from).")


def _load_investor_from_db(investor):
    """
    Load all historical quarters for one investor from 13f.db.
    Returns the same dict shape that fetch_investor_data used to return,
    with an extra 'holdingsHistory' list covering ALL available quarters.
    """
    inv_id = investor["id"]

    conn = _db_connect()

    # All filings newest-first
    filings = conn.execute("""
        SELECT id, accession, filing_date, quarter, holding_count, aum_millions
        FROM   filings
        WHERE  investor_id = ?
        ORDER  BY filing_date DESC
    """, (inv_id,)).fetchall()

    if not filings:
        conn.close()
        return None

    def load_holdings(filing_id):
        rows = conn.execute("""
            SELECT name, cusip, ticker, sector, value, shares, pct, put_call
            FROM   holdings
            WHERE  filing_id = ?
            ORDER  BY pct DESC
        """, (filing_id,)).fetchall()
        result = []
        for r in rows:
            result.append({
                "name":    r["name"],
                "cusip":   r["cusip"],
                "ticker":  corrected_ticker(r["ticker"], r["cusip"], r["name"]),
                "sector":  r["sector"],
                "value":   r["value"],
                "shares":  r["shares"],
                "pct":     r["pct"],
                "putCall": r["put_call"],
            })
        return result

    latest   = filings[0]
    holdings = load_holdings(latest["id"])

    prev_holdings = []
    prev_date     = None
    if len(filings) > 1:
        prev_holdings = load_holdings(filings[1]["id"])
        prev_date     = filings[1]["filing_date"]

    # Lightweight history metadata only. Full per-quarter holdings are served
    # on demand by /api/history_all, so we don't keep them in _all_data.
    history = []
    for f in filings[1:]:
        history.append({
            "quarter": f["quarter"],
            "date":    f["filing_date"],
            "aum":     f["aum_millions"],
        })

    conn.close()

    total_val = sum(h["value"] for h in holdings)
    aum_str   = f"${total_val/1000:.1f}B" if total_val >= 1000 else f"${total_val:.0f}M"

    result = {
        **investor,
        "aum":              aum_str,
        "aumRaw":           total_val,
        "latestQ":          latest["quarter"],
        "previousQ":        filings[1]["quarter"] if len(filings) > 1 else "",
        "filingDate":       latest["filing_date"],
        "lastFilingDate":   latest["filing_date"],
        "previousFilingDate": prev_date or "",
        "cacheRunDate":     date.today().isoformat(),
        "holdings":         holdings[:60],       # cards show top 10; keep a buffer for search
        "holdingsPrev":     prev_holdings[:60],
        "holdingsHistory":  history,
        "positionsAsOf":    latest["filing_date"],
        "portfolioPerfSinceFiling": None,   # enriched lazily below
    }

    # Enrich latest holdings with price performance (top 50)
    if _YF_AVAILABLE:
        # Price enrichment is DISABLED during bulk load. Calling enrich_performance
        # here (via yfinance) for all 47 investors was consuming gigabytes of RAM
        # and causing OOM. It stays None; it can be computed lazily per-investor.
        result["portfolioPerfSinceFiling"] = None

    return result

    


# ─────────────────────────────────────────────
#  MAIN FETCH FUNCTION
#  Tries DB first; falls back to live EDGAR fetch
#  if 13f.db does not exist yet (pre-ingest).
# ─────────────────────────────────────────────

def fetch_investor_data(investor):
    inv_id = investor["id"]
    cik    = investor["cik"]
    cp     = cache_path(inv_id, "all")

    # Fast path: read from SQLite DB (populated by ingest.py)
    if _db_available():
        print(f"  [{inv_id}] Loading from 13f.db ...")
        try:
            data = _load_investor_from_db(investor)
            if data:
                return data
            print(f"  [{inv_id}] Not in DB yet - falling back to EDGAR fetch")
        except Exception as e:
            print(f"  [{inv_id}] DB read error ({e}) - falling back to EDGAR fetch")

    # Slow path: live EDGAR fetch (used before ingest.py has been run)
    if cache_valid(cp):
        print(f"  [{inv_id}] Using cached data")
        cached = load_cache(cp)
        if cached and cached.get("portfolioPerfSinceFiling") is None and _YF_AVAILABLE:
            print(f"  [{inv_id}] Cache missing price data - enriching now...")
            portfolio_perf = enrich_performance(cached.get("holdings", []), cached.get("filingDate", ""))
            cached["portfolioPerfSinceFiling"] = portfolio_perf
            cached["positionsAsOf"] = cached.get("filingDate", "")
            save_cache(cp, cached)
        return cached

    print(f"  [{inv_id}] Fetching from SEC EDGAR (CIK {cik})...")
    filings = get_13f_filings(cik)

    if len(filings) < 1:
        print(f"  [{inv_id}] No 13F filings found")
        return None

    def fetch_holdings_for(filing):
        acc = filing["accession"]
        doc_url = get_filing_index(cik, acc)
        if not doc_url:
            print(f"  [{inv_id}] Could not find a holdings document for {acc}")
            return None, None
        print(f"  [{inv_id}] Downloading {doc_url}")
        doc = sec_get(doc_url)
        if not doc:
            return None, None
        # Auto-detects XML (post-2013) vs legacy fixed-width text (pre-2013).
        raw = parse_holdings(doc)
        enriched = enrich_with_tickers(raw)
        # Assign sectors (static + dynamic keyword)
        for h in enriched:
            h["sector"] = guess_sector(h["ticker"], h["name"])
        # Fill remaining "Other" via Yahoo Finance
        enrich_sectors_via_yahoo(enriched)
        return enriched, filing["date"]

    # Latest filing. If the newest filing cannot be parsed, try a few older ones
    # instead of failing the whole investor.
    holdings, latest_date = None, None
    latest_index = None

    for idx, filing in enumerate(filings[:2]):
        holdings, latest_date = fetch_holdings_for(filing)
        if holdings:
            latest_index = idx
            break
        print(f"  [{inv_id}] Could not load filing {filing['accession']}; trying older filing...")

    if not holdings:
        return None

    # Map a 13F FILING date to the REPORTING PERIOD it covers. Filings land up
    # to 45 days after quarter-end, so the filing's own calendar quarter is one
    # ahead of the period disclosed (a May filing reports Q1 holdings, not Q2).
    def date_to_quarter(d):
        if not d:
            return "Unknown"
        try:
            y, m, _ = d.split("-")
            y = int(y)
            filing_q = (int(m) - 1) // 3 + 1
            report_q = filing_q - 1
            if report_q == 0:        # filed Jan–Mar -> Q4 of prior year
                report_q = 4
                y -= 1
            return f"Q{report_q} {y}"
        except:
            return d

    # Fetch up to 3 historical quarters after the latest filing (Q-1, Q-2, Q-3).
    # We still keep holdingsPrev pointing at Q-1 so all existing QoQ logic is untouched.
    history = []   # list of {quarter, date, holdings} for Q-1 through Q-3
    if latest_index is not None:
        search_filings = filings[latest_index + 1:latest_index + 10]
        for filing in search_filings:
            if len(history) >= 3:
                break
            h, d = fetch_holdings_for(filing)
            if h:
                history.append({
                    "quarter":  date_to_quarter(d),
                    "date":     d,
                    "holdings": h,
                })

    # Backwards-compat: holdingsPrev + previousQ still point at Q-1
    prev_holdings = history[0]["holdings"] if history else []
    prev_date     = history[0]["date"]     if history else None

    # Calculate AUM
    total_val = sum(h["value"] for h in holdings)
    aum_str = f"${total_val/1000:.1f}B" if total_val >= 1000 else f"${total_val:.0f}M"

    # Cap positions based on fund type.
    # Concentrated funds (Buffett, Ackman, Burry): top 40 is fine.
    # Multi-strategy / quant funds (Citadel, Renaissance, Millennium,
    # Point72, D.E. Shaw, etc.) hold thousands — keep up to 500.
    result = {
        **investor,
        "aum":          aum_str,
        "aumRaw":       total_val,
        "latestQ":      date_to_quarter(latest_date),
        "previousQ":    date_to_quarter(prev_date) if prev_date else "",
        "filingDate":   latest_date or "",
        "lastFilingDate": latest_date or "",
        "previousFilingDate": prev_date or "",
        "cacheRunDate": date.today().isoformat(),
        "holdings":     holdings,
        "holdingsPrev": prev_holdings if prev_holdings else [],
        "holdingsHistory": history,
    }

    # Enrich top 50 holdings with price performance since filing
    portfolio_perf = enrich_performance(holdings, latest_date)

    result["portfolioPerfSinceFiling"] = portfolio_perf
    # Also store filing date clearly for the frontend lag indicator
    result["positionsAsOf"] = latest_date or ""

    save_cache(cp, result)
    print(f"  [{inv_id}] Done — {len(holdings)} holdings, AUM {aum_str}" +
          (f", portfolio perf: {portfolio_perf:+.1f}%" if portfolio_perf is not None else ""))
    return result


# ─────────────────────────────────────────────
#  HTTP SERVER
# ─────────────────────────────────────────────

_all_data = {}   # in-memory store populated at startup
_data_lock = threading.Lock()
# history_all cache — built once on first request, served from memory after that
# [0] = raw JSON bytes, [1] = gzip-compressed bytes
_history_cache      = [None, None]
_history_cache_lock = threading.Lock()



def clean_json_value(obj):
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    if isinstance(obj, dict):
        return {k: clean_json_value(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [clean_json_value(v) for v in obj]

    return obj









# ─────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT for backtest results
# ─────────────────────────────────────────────────────────────────────────
# Drop-in replacement for generate_backtest_xlsx in server.py
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

def generate_backtest_xlsx(bt_result):
    """Build a formula-driven .xlsx from a backtest result dict.

    Layout:
      • Summary  – all metrics as formulas (links to Quarters).
      • Backtest – combined constituents + dollar path, TRANSPOSED:
                   categories down the rows, each stock a column,
                   4 quarters across per band, then the next 4 wrap below.
                   Every derived number is a live Excel formula.
      • Quarters – per-quarter overview, derived columns link to Backtest.
    Returns the path to a temp .xlsx file.
    """
    A = "Arial"
    def Fn(**kw): return Font(name=A, **kw)
    BLUE=Fn(color="0000FF"); BLACK=Fn(color="000000"); GREEN=Fn(color="008000")
    LBL=Fn(bold=True, color="1F2937"); HEADW=Fn(bold=True, color="FFFFFF")
    TITLE=Fn(bold=True, size=14, color="1F2937"); SUBLBL=Fn(bold=True, size=9, color="6B7280")
    DARK=PatternFill("solid", fgColor="1F3A5F"); QFILL=PatternFill("solid", fgColor="2E5984")
    ROLLF=PatternFill("solid", fgColor="FBF3E2")
    thin=Side(style="thin", color="D0D7DE"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
    RIGHT=Alignment(horizontal="right"); CENTER=Alignment(horizontal="center")
    PX='#,##0.0000'; USD='$#,##0.00'; PCT='0.00%'; MULT='#,##0.0000"×"'; DT='yyyy-mm-dd'

    def to_date(s):
        try: return datetime.fromisoformat(str(s)[:19])
        except Exception: return s

    periods = bt_result.get("periods", [])
    n_q = len(periods)

    # ── band geometry: 4 quarters across, wrap below ─────────────────────
    BAND_QS = 4; GAP = 2; TOP = 5
    off = dict(qh=0, hf=1, ht=2, tk=3, ep=4, xp=5, ret=6, src=7, start=8, chg=9,
               end=10, drop=11, sec=12, n=13, qret=14, qstart=15, qend=16,
               cum=17, pv=18, cpl=19, qpl=20)
    BAND_H = max(off.values()) + 1
    labels = {
        'qh':"Quarter (held filing-date → next filing-date)", 'hf':"Held From", 'ht':"Held To",
        'tk':"Ticker", 'ep':"Entry Price", 'xp':"Exit Price", 'ret':"Return  =(Exit/Entry)−1",
        'src':"Price Source", 'start':"Start $  (equal slice = Q-Start ÷ N)",
        'chg':"Change $  =Slice × Return", 'end':"End $  =Slice + Change",
        'drop':"Dropped (no price)", 'sec':"── QUARTER ROLL-UP ──",
        'n':"Names priced (N)  =COUNT(returns)", 'qret':"Quarter Return  =AVERAGE(returns)",
        'qstart':"Quarter Start $  (= prior Quarter End $)", 'qend':"Quarter End $  =Start × (1+Q.Return)",
        'cum':"Cumulative Growth  =Q.End ÷ capital", 'pv':"Portfolio Value $  =Q.End $",
        'cpl':"Cumulative P&L $  =Value − capital", 'qpl':"Quarter P&L $  =Q.End − Q.Start",
    }

    wb = Workbook()

    # ============================================================ Backtest
    bt = wb.active; bt.title = "Backtest"; bt.sheet_properties.tabColor = "2E5984"
    bt["A1"] = "Equal-Weight Top-15 Consensus Backtest — wrapped 4 quarters per band (constituents + dollar path)"
    bt["A1"].font = TITLE
    bt["A2"] = "Starting capital ($)"; bt["A2"].font = LBL
    bt["B2"] = 100000; bt["B2"].font = BLUE; bt["B2"].number_format = '$#,##0'
    bt["A3"] = ("Categories down the rows; each stock a column; 4 quarters across per band, then the next 4 "
                "wrap below (Q1 2000 sits under Q1 1999). Blue = input. Black = formula.")
    bt["A3"].font = SUBLBL
    bt.column_dimensions['A'].width = 34

    def src_label(con):
        s0, s1 = con.get("src0"), con.get("src1")
        if s0 == "implied" and s1 == "implied": return "13F mark"
        if s0 == "implied" or s1 == "implied":  return "mixed"
        return "market"

    # per-quarter (firstcol, base_row)
    metas = []; col = 2
    for qi, p in enumerate(periods):
        pos = qi % BAND_QS; b = qi // BAND_QS
        base = TOP + b * (BAND_H + GAP)
        if pos == 0: col = 2
        cons = p.get("constituents", [])
        metas.append({'fc': col, 'base': base, 'cons': cons})
        col = col + max(len(cons), 1) + 1
    n_bands = (n_q + BAND_QS - 1) // BAND_QS if n_q else 0
    max_col = max((m['fc'] + max(len(m['cons']), 1) for m in metas), default=2)

    # row labels per band
    for b in range(n_bands):
        base = TOP + b * (BAND_H + GAP)
        for k, lab in labels.items():
            c = bt.cell(base + off[k], 1, lab)
            c.font = SUBLBL if k == 'sec' else LBL
            if k == 'sec': c.fill = ROLLF

    # blocks
    for qi, p in enumerate(periods):
        m = metas[qi]; fc = m['fc']; base = m['base']; cons = m['cons']
        fcL = get_column_letter(fc)
        def R(k): return base + off[k]
        last = fc + max(len(cons), 1) - 1
        bt.merge_cells(start_row=R('qh'), start_column=fc, end_row=R('qh'), end_column=last)
        h = bt.cell(R('qh'), fc, f"{p['from_quarter']} → {p['to_quarter']}")
        h.font = HEADW; h.fill = QFILL; h.alignment = CENTER
        bt.cell(R('hf'), fc, to_date(p['from'])).font = BLUE; bt.cell(R('hf'), fc).number_format = DT
        bt.cell(R('ht'), fc, to_date(p['to'])).font = BLUE; bt.cell(R('ht'), fc).number_format = DT

        for j, con in enumerate(cons):
            cc = fc + j; L = get_column_letter(cc)
            bt.cell(R('tk'), cc, con['ticker']).font = Fn(bold=True)
            e = bt.cell(R('ep'), cc, con['p0']); e.font = BLUE; e.number_format = PX
            x = bt.cell(R('xp'), cc, con['p1']); x.font = BLUE; x.number_format = PX
            rt = bt.cell(R('ret'), cc, f"=IF({L}{R('ep')}=0,0,({L}{R('xp')}/{L}{R('ep')})-1)")
            rt.font = BLACK; rt.number_format = PCT
            bt.cell(R('src'), cc, src_label(con)).font = Fn(size=9)
            st = bt.cell(R('start'), cc, f"=${fcL}${R('qstart')}/${fcL}${R('n')}"); st.font = BLACK; st.number_format = USD
            ch = bt.cell(R('chg'), cc, f"={L}{R('start')}*{L}{R('ret')}"); ch.font = BLACK; ch.number_format = USD
            en = bt.cell(R('end'), cc, f"={L}{R('start')}+{L}{R('chg')}"); en.font = BLACK; en.number_format = USD
            for rk in ('tk','ep','xp','ret','src','start','chg','end'): bt.cell(R(rk), cc).alignment = RIGHT
            bt.column_dimensions[L].width = 11

        drp = p.get("dropped", [])
        if drp: bt.cell(R('drop'), fc, "; ".join(drp)).font = Fn(size=9, color="B45309")

        if len(cons):
            f0 = get_column_letter(fc); fL = get_column_letter(fc + len(cons) - 1)
            rng = f"{f0}{R('ret')}:{fL}{R('ret')}"
            bt.cell(R('n'), fc, f"=COUNT({rng})")
            bt.cell(R('qret'), fc, f"=AVERAGE({rng})")
        else:
            bt.cell(R('n'), fc, 0); bt.cell(R('qret'), fc, 0)
        if qi == 0:
            bt.cell(R('qstart'), fc, "=$B$2")
        else:
            pm = metas[qi-1]; pL = get_column_letter(pm['fc'])
            bt.cell(R('qstart'), fc, f"={pL}{pm['base']+off['qend']}")
        bt.cell(R('qend'), fc, f"={fcL}{R('qstart')}*(1+{fcL}{R('qret')})")
        bt.cell(R('cum'),  fc, f"={fcL}{R('qend')}/$B$2")
        bt.cell(R('pv'),   fc, f"={fcL}{R('qend')}")
        bt.cell(R('cpl'),  fc, f"={fcL}{R('pv')}-$B$2")
        bt.cell(R('qpl'),  fc, f"={fcL}{R('qend')}-{fcL}{R('qstart')}")
        for k, fmt in [('n','0'),('qret',PCT),('qstart',USD),('qend',USD),
                       ('cum',MULT),('pv',USD),('cpl',USD),('qpl',USD)]:
            c = bt.cell(R(k), fc); c.font = BLACK; c.number_format = fmt; c.alignment = RIGHT; c.fill = ROLLF

    # dollar-path table + chart below all bands
    base_path = TOP + n_bands * (BAND_H + GAP) + 2
    bt.cell(base_path-1, 1, "DOLLAR PATH OVER TIME  (links to each quarter's roll-up above)").font = SUBLBL
    for j, hh in enumerate(["Date","Cumulative Growth (×)","Portfolio Value $","Cumulative P&L $"]):
        c = bt.cell(base_path, 1+j, hh); c.font = HEADW; c.fill = DARK; c.alignment = CENTER; c.border = BORD
    if n_q:
        r0 = base_path + 1; m0 = metas[0]; fc0 = get_column_letter(m0['fc'])
        bt.cell(r0,1, f"={fc0}{m0['base']+off['hf']}"); bt.cell(r0,1).number_format=DT; bt.cell(r0,1).font=GREEN
        bt.cell(r0,3, "=$B$2"); bt.cell(r0,3).number_format=USD; bt.cell(r0,3).font=BLACK
        bt.cell(r0,2, f"=C{r0}/$B$2"); bt.cell(r0,2).number_format=MULT; bt.cell(r0,2).font=BLACK
        bt.cell(r0,4, f"=C{r0}-$B$2"); bt.cell(r0,4).number_format=USD; bt.cell(r0,4).font=BLACK
        for qi in range(n_q):
            rr = base_path + 2 + qi; m = metas[qi]; L = get_column_letter(m['fc'])
            bt.cell(rr,1, f"={L}{m['base']+off['ht']}"); bt.cell(rr,1).number_format=DT; bt.cell(rr,1).font=GREEN
            bt.cell(rr,2, f"={L}{m['base']+off['cum']}"); bt.cell(rr,2).number_format=MULT; bt.cell(rr,2).font=GREEN
            bt.cell(rr,3, f"={L}{m['base']+off['pv']}");  bt.cell(rr,3).number_format=USD;  bt.cell(rr,3).font=GREEN
            bt.cell(rr,4, f"={L}{m['base']+off['cpl']}"); bt.cell(rr,4).number_format=USD;  bt.cell(rr,4).font=GREEN
            for cc in (1,2,3,4): bt.cell(rr,cc).border = BORD
        last_path = base_path + 1 + n_q
        for w,c in [(16,'B'),(16,'C'),(16,'D')]: bt.column_dimensions[c].width = w
        chart = LineChart(); chart.title = "Portfolio Value Over Time (equity curve / cumulative P&L)"
        chart.style = 12; chart.height = 9; chart.width = 26
        chart.y_axis.title = "Portfolio value ($)"; chart.x_axis.title = "Date"
        chart.x_axis.number_format = 'yyyy'; chart.x_axis.majorTimeUnit = "years"
        chart.add_data(Reference(bt, min_col=3, min_row=base_path, max_row=last_path), titles_from_data=True)
        chart.set_categories(Reference(bt, min_col=1, min_row=base_path+1, max_row=last_path))
        chart.series[0].graphicalProperties.line.width = 22000
        bt.add_chart(chart, f"F{base_path}")

    # ============================================================ Quarters
    qs = wb.create_sheet("Quarters"); qs.sheet_properties.tabColor = "0D7D5F"
    q_headers = ["From Quarter","To Quarter","Held From","Held To","Names In","Names Priced",
                 "Implied Legs","Dropped","Quarter Return","Cumulative","Quarter Start $",
                 "Quarter End $","P&L $","Investors","Coverage"]
    for j, hh in enumerate(q_headers):
        c = qs.cell(1, 1+j, hh); c.font = HEADW; c.fill = DARK; c.alignment = CENTER; c.border = BORD
    for qi, p in enumerate(periods):
        r = qi + 2; m = metas[qi]; L = get_column_letter(m['fc']); base = m['base']
        qs.cell(r,1, p['from_quarter']); qs.cell(r,2, p['to_quarter'])
        qs.cell(r,3, to_date(p['from'])).number_format = DT; qs.cell(r,4, to_date(p['to'])).number_format = DT
        qs.cell(r,3).font = BLUE; qs.cell(r,4).font = BLUE
        qs.cell(r,5, p['n_in']).font = BLUE
        qs.cell(r,6, f"=Backtest!{L}{base+off['n']}").font = GREEN
        qs.cell(r,7, p.get('n_implied', 0)).font = BLUE
        qs.cell(r,8, ", ".join(p.get('dropped', [])))
        qs.cell(r,9, f"=Backtest!{L}{base+off['qret']}").font = GREEN;  qs.cell(r,9).number_format = PCT
        qs.cell(r,10,f"=Backtest!{L}{base+off['cum']}").font = GREEN;   qs.cell(r,10).number_format = MULT
        qs.cell(r,11,f"=Backtest!{L}{base+off['qstart']}").font = GREEN;qs.cell(r,11).number_format = USD
        qs.cell(r,12,f"=Backtest!{L}{base+off['qend']}").font = GREEN;  qs.cell(r,12).number_format = USD
        qs.cell(r,13,f"=Backtest!{L}{base+off['qpl']}").font = GREEN;   qs.cell(r,13).number_format = USD
        qs.cell(r,14,"; ".join(p.get('investors', []))).font = Fn(size=8)
        qs.cell(r,15,f"=IF(E{r}=0,0,F{r}/E{r})").font = BLACK; qs.cell(r,15).number_format = PCT
    last_q = n_q + 1
    for c,w in {'A':10,'B':10,'C':12,'D':12,'E':9,'F':9,'G':9,'H':10,'I':11,'J':12,
                'K':14,'L':14,'M':14,'N':60,'O':10}.items():
        qs.column_dimensions[c].width = w
    qs.freeze_panes = "A2"

    # ============================================================ Summary
    sm = wb.create_sheet("Summary"); sm.sheet_properties.tabColor = "1549A8"
    wb.move_sheet("Summary", -(wb.sheetnames.index("Summary")))
    sm["A1"] = "Backtest Summary"; sm["A1"].font = TITLE
    rows = [
        ("Start date",           f"=Quarters!C2", DT, GREEN),
        ("End date",             f"=Quarters!D{last_q}", DT, GREEN),
        ("Starting capital ($)", f"=Backtest!B2", '$#,##0', GREEN),
        ("Total return",         f"=Quarters!J{last_q}-1", PCT, GREEN),
        ("CAGR",                 f"=IF(Quarters!D{last_q}=Quarters!C2,0,Quarters!J{last_q}^(365.25/(Quarters!D{last_q}-Quarters!C2))-1)", PCT, GREEN),
        ("Ending value ($)",     f"=Backtest!B2*Quarters!J{last_q}", USD, GREEN),
        ("Net P&L ($)",          f"=B7-Backtest!B2", USD, BLACK),
        ("Quarters",             f"=COUNTA(Quarters!A2:A{last_q})", '0', GREEN),
        ("Avg coverage",         f"=AVERAGE(Quarters!O2:O{last_q})", PCT, GREEN),
        ("Implied price legs",   f"=SUM(Quarters!G2:G{last_q})", '0', GREEN),
    ]
    for i,(lab,formula,fmt,fnt) in enumerate(rows):
        r = i + 2; sm.cell(r,1,lab).font = LBL
        c = sm.cell(r,2,formula); c.font = fnt; c.number_format = fmt; c.alignment = RIGHT
    sm.column_dimensions['A'].width = 22; sm.column_dimensions['B'].width = 18

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(CACHE_DIR))
    wb.save(tmp.name); tmp.close()
    return tmp.name


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
         super().__init__(*args, directory=str(Path(__file__).parent), **kwargs)

    def _send_json(self, payload_bytes, status=200):
        """
        Send a JSON response, using gzip compression if the client supports it.
        Sets Content-Length so the connection never hangs waiting for EOF.
        Silently absorbs BrokenPipeError / ConnectionResetError so a browser
        closing the tab early never crashes the server thread.
        """
        accept_enc = self.headers.get("Accept-Encoding", "")
        use_gzip = "gzip" in accept_enc

        if use_gzip:
            buf = io.BytesIO()
            with gzip.GzipFile(fileobj=buf, mode="wb", compresslevel=6) as gz:
                gz.write(payload_bytes)
            body = buf.getvalue()
        else:
            body = payload_bytes

        try:
            self.send_response(status)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Content-Length", str(len(body)))
            if use_gzip:
                self.send_header("Content-Encoding", "gzip")
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError):
            pass  # client disconnected — nothing we can do

    def do_POST(self):
        try:
            if self.path == "/api/backtest":
                length = int(self.headers.get("Content-Length", 0) or 0)
                raw = self.rfile.read(length) if length else b"{}"

                try:
                    req = json.loads(raw.decode("utf-8") or "{}")
                except Exception:
                    self._send_json(b'{"ok":false,"error":"Invalid JSON body"}', status=400)
                    return

                print(f"  [API] /api/backtest: {len(req.get('baskets', []))} baskets")

                result = run_consensus_backtest(
                    req.get("baskets", []),
                    include_partial=bool(req.get("includePartial", False)),
                )

                result = clean_json_value(result)
                payload = json.dumps(result, default=str, allow_nan=False).encode()
                self._send_json(payload)
                return
            if self.path == "/api/backtest/xlsx":
                if not _OPENPYXL_AVAILABLE:
                    self._send_json(b'{"ok":false,"error":"openpyxl not installed. Run: pip install openpyxl"}', status=500)
                    return
                length = int(self.headers.get("Content-Length", 0) or 0)
                raw = self.rfile.read(length) if length else b"{}"
                try:
                    bt_result = json.loads(raw.decode("utf-8") or "{}")
                except Exception:
                    self._send_json(b'{"ok":false,"error":"Invalid JSON body"}', status=400)
                    return
                print(f"  [API] /api/backtest/xlsx: generating Excel export")
                try:
                    xlsx_path = generate_backtest_xlsx(bt_result)
                    with open(xlsx_path, "rb") as f:
                        data = f.read()
                    os.unlink(xlsx_path)
                    self.send_response(200)
                    self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.send_header("Content-Disposition", 'attachment; filename="consensus_backtest.xlsx"')
                    self.send_header("Content-Length", str(len(data)))
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.end_headers()
                    self.wfile.write(data)
                except Exception as e:
                    print(f"  [API] xlsx error: {e}")
                    self._send_json(json.dumps({"ok": False, "error": str(e)}).encode(), status=500)
                return
            if self.path == "/api/optimize":
                length = int(self.headers.get("Content-Length", 0) or 0)
                raw = self.rfile.read(length) if length else b"{}"

                try:
                    req = json.loads(raw.decode("utf-8") or "{}")
                except Exception:
                    self._send_json(b'{"ok":false,"error":"Invalid JSON body"}', status=400)
                    return

                print(f"  [API] /api/optimize: pool={len(req.get('investorIds', []))} "
                    f"method={req.get('method')} budget={req.get('budget')}")

                result = run_cagr_optimizer(req)

                print(f"  [API] /api/optimize done: {result.get('evaluations')} evals, "
                    f"{len(result.get('results', []))} ranked")

                result = clean_json_value(result)
                payload = json.dumps(result, default=str, allow_nan=False).encode()
                self._send_json(payload)
                return
            self._send_json(b'{"ok":false,"error":"Unknown endpoint"}', status=404)
        except (BrokenPipeError, ConnectionResetError):
            pass
        except Exception as e:
            print(f"  [API] POST error: {e}")
            try:
                self._send_json(json.dumps({"ok": False, "error": str(e)}).encode(), status=500)
            except Exception:
                pass

    def do_GET(self):
        try:
            self._do_GET_inner()
        except (BrokenPipeError, ConnectionResetError):
            pass  # client closed connection early — suppress cleanly

    def _do_GET_inner(self):
        # ── /api/data  ────────────────────────────────────────────────────────
        if self.path == "/api/data":
            with _data_lock:
                # Build a lean summary WITHOUT deep-copying the whole dataset.
                # Only the fields the investor grid actually renders are included;
                # heavy per-quarter holdings are served by /api/history_all instead.
                summary = []
                for inv in _all_data.values():
                    summary.append({
                        "id":                inv.get("id"),
                        "name":              inv.get("name"),
                        "firm":              inv.get("firm"),
                        "strategy":          inv.get("strategy"),
                        "category":          inv.get("category"),
                        "color":             inv.get("color"),
                        "cik":               inv.get("cik"),
                        "aum":               inv.get("aum"),
                        "aumRaw":            inv.get("aumRaw"),
                        "latestQ":           inv.get("latestQ"),
                        "previousQ":         inv.get("previousQ"),
                        "filingDate":        inv.get("filingDate"),
                        "lastFilingDate":    inv.get("lastFilingDate"),
                        "previousFilingDate": inv.get("previousFilingDate"),
                        "positionsAsOf":     inv.get("positionsAsOf"),
                        "portfolioPerfSinceFiling": inv.get("portfolioPerfSinceFiling"),
                        "holdings":          inv.get("holdings", [])[:60],
                        "holdingsPrev":      inv.get("holdingsPrev", [])[:60],
                    })

            summary = clean_json_value(summary)
            payload = json.dumps(summary, default=str, allow_nan=False).encode()
            self._send_json(payload)
            return
        # ── /api/enrich/<investor_id>  ────────────────────────────────────────
        # Lazy, on-demand price + fundamentals enrichment for ONE investor.
        # Enrichment (yfinance) is deliberately NOT run at startup because doing
        # it for all 47 investors at once exhausted memory. Instead the frontend
        # calls this when an investor is opened, so we only ever enrich the few
        # investors actually viewed. Result is cached on the investor dict, so a
        # second view is instant and costs no extra memory/network.
        if self.path.startswith("/api/enrich/"):
            inv_id = self.path[len("/api/enrich/"):].strip("/")
            with _data_lock:
                inv = _all_data.get(inv_id)
            if not inv:
                self._send_json(b'{"ok":false,"error":"Unknown investor"}', status=404)
                return

            # Already enriched this session? Serve the cached holdings straight back.
            if inv.get("_enriched"):
                out = clean_json_value({
                    "ok": True,
                    "id": inv_id,
                    "portfolioPerfSinceFiling": inv.get("portfolioPerfSinceFiling"),
                    "holdings": inv.get("holdings", []),
                })
                self._send_json(json.dumps(out, default=str, allow_nan=False).encode())
                return

            if not _YF_AVAILABLE:
                self._send_json(b'{"ok":false,"error":"Price data unavailable on server"}', status=503)
                return

            try:
                # enrich_performance mutates the holding dicts in place (adds
                # priceAtFiling/currentPrice/perfSinceFiling/fundamentals) and
                # returns the weighted portfolio performance.
                holdings = inv.get("holdings", [])
                perf = enrich_performance(holdings, inv.get("filingDate", ""))
                with _data_lock:
                    inv["portfolioPerfSinceFiling"] = perf
                    inv["_enriched"] = True
                out = {
                    "ok": True,
                    "id": inv_id,
                    "portfolioPerfSinceFiling": perf,
                    "holdings": holdings,
                }
                out = clean_json_value(out)
                self._send_json(json.dumps(out, default=str, allow_nan=False).encode())
            except Exception as e:
                self._send_json(
                    json.dumps({"ok": False, "error": str(e)}).encode(), status=500)
            return

        # ── /api/history_all  ─────────────────────────────────────────────────
        # This endpoint returns ALL historical holdings for ALL investors.
        # The raw JSON can be 10-50 MB, so we:
        #   1. Build it once and cache the raw bytes in memory (_history_cache)
        #   2. Gzip it once and cache the compressed bytes (_history_cache_gz)
        #   3. Serve from memory on every subsequent request — near-instant
        if self.path == "/api/history_all":
            if not _db_available():
                self._send_json(b'{"error":"DB not available"}', status=503)
                return
            try:
                with _history_cache_lock:
                    if _history_cache[0] is None:
                        # First request — build from DB and cache
                        print("  [API] Building history_all cache from DB...")
                        t0 = time.time()
                        conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
                        conn.row_factory = sqlite3.Row

                        # Query 1: all filings
                        filings = conn.execute("""
                            SELECT id, investor_id, filing_date, quarter, aum_millions
                            FROM   filings
                            ORDER  BY investor_id, filing_date DESC
                        """).fetchall()

                        # Query 2: ALL holdings in one shot, keyed by filing_id
                        hrows = conn.execute("""
                            SELECT filing_id, ticker, name, sector,
                                   value, shares, pct, put_call
                            FROM   holdings
                            ORDER  BY filing_id, pct DESC
                        """).fetchall()
                        conn.close()

                        # Build filing_id → holdings list
                        hmap = {}
                        for h in hrows:
                            fid = h["filing_id"]
                            if fid not in hmap:
                                hmap[fid] = []
                            hmap[fid].append({
                                "ticker":   h["ticker"],
                                "name":     h["name"],
                                "sector":   h["sector"],
                                "value":    h["value"],
                                "shares":   h["shares"],
                                "pct":      h["pct"],
                                "put_call": h["put_call"],
                            })

                        # Assemble result grouped by investor
                        result = {}
                        for f in filings:
                            inv_id = f["investor_id"]
                            if inv_id not in result:
                                result[inv_id] = []
                            result[inv_id].append({
                                "quarter":  f["quarter"],
                                "date":     f["filing_date"],
                                "aum":      f["aum_millions"],
                                "holdings": hmap.get(f["id"], []),
                            })

                        raw = json.dumps(result, default=str).encode()
                        buf = io.BytesIO()
                        with gzip.GzipFile(fileobj=buf, mode="wb", compresslevel=6) as gz:
                            gz.write(raw)
                        _history_cache[0] = raw
                        _history_cache[1] = buf.getvalue()
                        elapsed = time.time() - t0
                        print(f"  [API] history_all cache built: "
                              f"{len(raw)//1024}KB raw / "
                              f"{len(_history_cache[1])//1024}KB gzip "
                              f"in {elapsed:.1f}s")

                # Serve from cache — check if client accepts gzip
                accept_enc = self.headers.get("Accept-Encoding", "")
                body   = _history_cache[1] if "gzip" in accept_enc else _history_cache[0]
                use_gz = "gzip" in accept_enc
                try:
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json")
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.send_header("Content-Length", str(len(body)))
                    if use_gz:
                        self.send_header("Content-Encoding", "gzip")
                    self.end_headers()
                    self.wfile.write(body)
                except (BrokenPipeError, ConnectionResetError):
                    pass

            except (BrokenPipeError, ConnectionResetError):
                pass
            except Exception as e:
                print(f"  [API] history_all error: {e}")
                try:
                    self._send_json(json.dumps({"error": str(e)}).encode(), status=500)
                except Exception:
                    pass
            return

        # ── /api/refresh/<id>  ────────────────────────────────────────────────
        if self.path.startswith("/api/refresh/"):
            inv_id = self.path.split("/")[-1]
            inv = next((i for i in INVESTORS if i["id"] == inv_id), None)

            if not inv:
                self._send_json(b'{"ok":false,"error":"Unknown investor"}', status=404)
                return

            cp = cache_path(inv_id, "all")
            if cp.exists():
                cp.unlink()

            data = fetch_investor_data(inv)
            if data:
                with _data_lock:
                    _all_data[inv_id] = data

            with _data_lock:
                refreshed = copy.deepcopy(_all_data.get(inv_id, {}))

            refreshed = clean_json_value(refreshed)

            payload = json.dumps(refreshed, default=str, allow_nan=False).encode()
            self._send_json(payload)
            return

        # ── Static files  ─────────────────────────────────────────────────────
        super().do_GET()

    def log_message(self, format, *args):
        # Suppress default logging noise for static files
        if "/api/" in str(args):
            print(f"  [API] {args[0]}")


def load_all_investors():
    """Load all investors - from DB if available, else live EDGAR fetch."""
    delay = 0.0 if _db_available() else 1.0
    for inv in INVESTORS:
        data = fetch_investor_data(inv)
        if data:
            with _data_lock:
                _all_data[inv["id"]] = data
        if delay:
            time.sleep(delay)
    src = "13f.db" if _db_available() else "SEC EDGAR"
    print(f"\n  Loaded {len(_all_data)}/{len(INVESTORS)} investors from {src}\n")


if __name__ == "__main__":
    PORT = int(os.environ.get("PORT", 8002))
    print("=" * 55)
    print("  13F Tracker - Local Server")

    print("=" * 55)

    # If 13f.db is missing but the compressed dump exists, rebuild it
    _rebuild_db_from_dump()

    if _db_available():
        size_mb = DB_PATH.stat().st_size / 1_000_000
        print(f"\n  Reading from 13f.db ({size_mb:.0f} MB) ...")
    else:
        print("\n  13f.db not found (and no 13f.sql.gz to rebuild from).")
        print("  Falling back to live SEC EDGAR fetch (slow).")
        print("  (Cached for 24h - delete /cache to force refresh)\n")

    print("\n  Step 1/2: Loading SEC ticker + sector database...")
    load_sec_ticker_file()
    _load_yf_cache()
    _load_price_cache()

    print("\n  Step 2/2: Loading investor data...")
    threading.Thread(target=load_all_investors, daemon=True).start()
    threading.Thread(target=load_all_investors, daemon=True).start()

    def _delayed_enrich():
        # wait until the base data is loaded, then enrich slowly in the background
        while len(_all_data) < len(INVESTORS):
            time.sleep(2)
        time.sleep(5)   # let startup settle fully first
        background_enrich_all(batch_size=5, pause_sec=3)
    threading.Thread(target=_delayed_enrich, daemon=True).start()

    print(f"  Server running at http://localhost:{PORT}")
    print(f"  Press Ctrl+C to stop\n")

    try:
        server = http.server.ThreadingHTTPServer(("", PORT), Handler)
    except OSError as e:
        if getattr(e, "errno", None) == 48:
            PORT = 8001
            print(f"  Port 8000 was busy, using http://localhost:{PORT} instead")
            server = http.server.ThreadingHTTPServer(("", PORT), Handler)
        else:
            raise
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Server stopped.")
import resource
mem = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024   # Linux: KB→MB
print(f"  Peak memory after full load: {mem:.0f} MB", flush=True)